"""Importación masiva de clientes y proveedores desde una planilla Excel.

A diferencia del cruce QMS/Defontana, aquí no hay un sistema externo con su
propio formato: la planilla es la que nosotros mismos generamos como plantilla
(ver plantilla_clientes / plantilla_proveedores), así que el encabezado se
busca por nombre de columna en cualquiera de las primeras filas, tolerando que
el usuario reordene o borre columnas que no le interesan.
"""

import unicodedata

from openpyxl import load_workbook

from app.extensions import db
from app.models.cliente import Cliente, Proveedor
from app.utils.rut import es_rut_valido, formatear_rut

COLUMNAS_CLIENTES = ("RUT", "Razón social", "Giro", "Dirección", "Comuna", "Ciudad", "Teléfono", "Email", "Contacto", "Activo")
COLUMNAS_PROVEEDORES = ("RUT", "Razón social", "Giro", "Dirección", "Teléfono", "Email", "Contacto", "Activo")


def _sin_acentos(texto: str) -> str:
    normalizado = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in normalizado if not unicodedata.combining(c))


def _clave(texto: str) -> str:
    return _sin_acentos((texto or "").strip().lower())


def _texto(valor, limite):
    if valor is None:
        return ""
    return str(valor).strip()[:limite]


def _es_activo(valor) -> bool:
    if valor is None or valor == "":
        return True
    texto = _clave(str(valor))
    return texto not in ("no", "0", "false", "inactivo", "n")


def _fila_encabezado(hoja):
    """Busca, en las primeras filas, la que tiene 'rut' en alguna celda."""
    for numero_fila in range(1, min(hoja.max_row, 15) + 1):
        for celda in hoja[numero_fila]:
            if celda.value and _clave(str(celda.value)) == "rut":
                return numero_fila
    return None


def _mapa_columnas(hoja, fila_encabezado, columnas_esperadas):
    """{nombre_de_columna_esperado: índice} según lo que encuentre en la fila de encabezado."""
    indices = {}
    for celda in hoja[fila_encabezado]:
        if not celda.value:
            continue
        titulo = _clave(str(celda.value))
        for esperado in columnas_esperadas:
            if _clave(esperado) == titulo:
                indices[esperado] = celda.column
    return indices


def _leer_filas(file_storage, columnas_esperadas):
    libro = load_workbook(file_storage.stream, data_only=True, read_only=True)
    hoja = libro.active

    fila_encabezado = _fila_encabezado(hoja)
    if fila_encabezado is None:
        raise ValueError("No se encontró una columna 'RUT' en el archivo. Usa la plantilla descargable.")

    indices = _mapa_columnas(hoja, fila_encabezado, columnas_esperadas)
    if "RUT" not in indices:
        raise ValueError("No se encontró la columna 'RUT'.")

    def _valor(fila, columna):
        indice = indices.get(columna)
        if not indice:
            return None
        return fila[indice - 1].value

    filas = []
    for fila in hoja.iter_rows(min_row=fila_encabezado + 1):
        rut_crudo = _valor(fila, "RUT")
        if rut_crudo is None or str(rut_crudo).strip() == "":
            continue
        filas.append((fila, _valor))
    return filas, _valor


def importar_clientes(file_storage, empresa_id: int) -> dict:
    filas, _valor = _leer_filas(file_storage, COLUMNAS_CLIENTES)

    existentes = {c.rut: c for c in Cliente.query.filter_by(empresa_id=empresa_id).all()}
    creados = 0
    actualizados = 0
    invalidos = []
    nuevos = []

    for fila, valor in filas:
        rut_crudo = str(valor(fila, "RUT")).strip()
        if not es_rut_valido(rut_crudo):
            invalidos.append(rut_crudo)
            continue
        rut = formatear_rut(rut_crudo)

        datos = {
            "razon_social": _texto(valor(fila, "Razón social"), 150),
            "giro": _texto(valor(fila, "Giro"), 150) or None,
            "direccion": _texto(valor(fila, "Dirección"), 200) or None,
            "comuna": _texto(valor(fila, "Comuna"), 80) or None,
            "ciudad": _texto(valor(fila, "Ciudad"), 80) or None,
            "telefono": _texto(valor(fila, "Teléfono"), 30) or None,
            "email": _texto(valor(fila, "Email"), 120) or None,
            "contacto_nombre": _texto(valor(fila, "Contacto"), 120) or None,
            "activo": _es_activo(valor(fila, "Activo")),
        }
        if not datos["razon_social"]:
            invalidos.append(rut_crudo)
            continue

        cliente = existentes.get(rut)
        if cliente is None:
            cliente = Cliente(empresa_id=empresa_id, rut=rut, **datos)
            nuevos.append(cliente)
            existentes[rut] = cliente
            creados += 1
        else:
            for campo, dato in datos.items():
                setattr(cliente, campo, dato)
            actualizados += 1

    if nuevos:
        db.session.add_all(nuevos)
    db.session.commit()
    return {"creados": creados, "actualizados": actualizados, "invalidos": invalidos}


def importar_proveedores(file_storage, empresa_id: int) -> dict:
    filas, _valor = _leer_filas(file_storage, COLUMNAS_PROVEEDORES)

    existentes = {p.rut: p for p in Proveedor.query.filter_by(empresa_id=empresa_id).all()}
    creados = 0
    actualizados = 0
    invalidos = []
    nuevos = []

    for fila, valor in filas:
        rut_crudo = str(valor(fila, "RUT")).strip()
        if not es_rut_valido(rut_crudo):
            invalidos.append(rut_crudo)
            continue
        rut = formatear_rut(rut_crudo)

        datos = {
            "razon_social": _texto(valor(fila, "Razón social"), 150),
            "giro": _texto(valor(fila, "Giro"), 150) or None,
            "direccion": _texto(valor(fila, "Dirección"), 200) or None,
            "telefono": _texto(valor(fila, "Teléfono"), 30) or None,
            "email": _texto(valor(fila, "Email"), 120) or None,
            "contacto_nombre": _texto(valor(fila, "Contacto"), 120) or None,
            "activo": _es_activo(valor(fila, "Activo")),
        }
        if not datos["razon_social"]:
            invalidos.append(rut_crudo)
            continue

        proveedor = existentes.get(rut)
        if proveedor is None:
            proveedor = Proveedor(empresa_id=empresa_id, rut=rut, **datos)
            nuevos.append(proveedor)
            existentes[rut] = proveedor
            creados += 1
        else:
            for campo, dato in datos.items():
                setattr(proveedor, campo, dato)
            actualizados += 1

    if nuevos:
        db.session.add_all(nuevos)
    db.session.commit()
    return {"creados": creados, "actualizados": actualizados, "invalidos": invalidos}
