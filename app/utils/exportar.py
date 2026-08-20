"""Generación de informes en Excel (.xlsx).

Todas las descargas de la aplicación pasan por aquí para que salgan con el mismo
formato: encabezado de marca, títulos fijos al hacer scroll, filtros activados,
anchos de columna razonables y totales al pie.

Uso típico desde una vista:

    columnas = [
        col("Código", ancho=18),
        col("Stock", ancho=10, formato=ENTERO, total="suma"),
        col("Valor", ancho=16, formato=CLP, total="suma"),
    ]
    filas = [[i.codigo, i.cantidad, i.valor] for i in items]
    return responder_excel("stock", "Stock y conteo", columnas, filas)
"""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from flask import make_response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Formatos de número más usados
CLP = '"$"#,##0;[Red]-"$"#,##0'
ENTERO = "#,##0"
DECIMAL = "#,##0.00"
# Cantidades de inventario: los decimales se muestran sólo si existen, para que
# un stock de 1.234 unidades no salga como "1.234,000" al lado de "12,5 metros".
CANTIDAD = "#,##0.###"
FECHA = "DD-MM-YYYY"
PORCENTAJE = "0.0%"

# Colores de marca (los mismos del CSS)
_AZUL = "0E2C47"
_AZUL_SUAVE = "DCEAF4"
_GRIS_LINEA = "E3E9F0"
_GRIS_TEXTO = "6B7A8D"


def col(titulo, ancho=16, formato=None, total=None):
    """Define una columna del informe.

    total: 'suma' para totalizar al pie, 'texto' para poner una etiqueta, None para nada.
    """
    return {"titulo": titulo, "ancho": ancho, "formato": formato, "total": total}


def _escribir_encabezado(hoja, titulo, subtitulo, cantidad_columnas):
    """Bloque de marca con el nombre del informe, la fecha y los filtros aplicados."""
    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, cantidad_columnas))
    celda = hoja.cell(row=1, column=1, value=f"Almex Operaciones — {titulo}")
    celda.font = Font(bold=True, size=14, color=_AZUL)
    celda.alignment = Alignment(vertical="center")
    hoja.row_dimensions[1].height = 22

    detalle = f"Generado el {date.today().strftime('%d-%m-%Y')}"
    if subtitulo:
        detalle = f"{subtitulo} · {detalle}"
    hoja.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, cantidad_columnas))
    celda = hoja.cell(row=2, column=1, value=detalle)
    celda.font = Font(size=9, color=_GRIS_TEXTO)


def _escribir_titulos(hoja, columnas, fila_titulos):
    borde_inferior = Border(bottom=Side(style="thin", color=_AZUL))
    for indice, columna in enumerate(columnas, start=1):
        celda = hoja.cell(row=fila_titulos, column=indice, value=columna["titulo"])
        celda.font = Font(bold=True, color="FFFFFF", size=10)
        celda.fill = PatternFill("solid", fgColor=_AZUL)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = borde_inferior
        hoja.column_dimensions[get_column_letter(indice)].width = columna["ancho"]
    hoja.row_dimensions[fila_titulos].height = 26


def _escribir_filas(hoja, columnas, filas, fila_inicio):
    borde = Border(bottom=Side(style="hair", color=_GRIS_LINEA))
    for numero, valores in enumerate(filas):
        fila_excel = fila_inicio + numero
        for indice, columna in enumerate(columnas, start=1):
            valor = valores[indice - 1] if indice - 1 < len(valores) else None
            if isinstance(valor, datetime):
                valor = valor.replace(tzinfo=None)
            celda = hoja.cell(row=fila_excel, column=indice, value=valor)
            celda.border = borde
            if columna["formato"]:
                celda.number_format = columna["formato"]
            if isinstance(valor, str) and len(valor) > 40:
                celda.alignment = Alignment(wrap_text=False)
        if numero % 2 == 1:  # bandeado suave para leer filas largas
            for indice in range(1, len(columnas) + 1):
                hoja.cell(row=fila_excel, column=indice).fill = PatternFill("solid", fgColor="FAFBFC")
    return fila_inicio + len(filas)


def _escribir_totales(hoja, columnas, filas, fila_totales):
    """Fila de totales con la suma real de cada columna marcada como 'suma'."""
    borde = Border(top=Side(style="thin", color=_AZUL))
    hay_totales = False

    for indice, columna in enumerate(columnas, start=1):
        celda = hoja.cell(row=fila_totales, column=indice)
        celda.border = borde
        celda.font = Font(bold=True, color=_AZUL)
        celda.fill = PatternFill("solid", fgColor=_AZUL_SUAVE)

        if columna["total"] == "suma":
            hay_totales = True
            total = sum(
                valores[indice - 1]
                for valores in filas
                if indice - 1 < len(valores)
                and isinstance(valores[indice - 1], (int, float, Decimal))
                and not isinstance(valores[indice - 1], bool)
            )
            celda.value = total
            if columna["formato"]:
                celda.number_format = columna["formato"]
        elif columna["total"] == "texto":
            hay_totales = True
            celda.value = f"Totales ({len(filas)})"

    return hay_totales


# Excel no acepta estos caracteres en el nombre de una hoja y revienta al
# guardar. Un título con una barra —"SII / Defontana"— es de lo más razonable
# para el informe, así que se limpia acá en vez de prohibirlo en cada pantalla.
_PROHIBIDOS_EN_HOJA = str.maketrans({c: "-" for c in r":\/?*[]"})


def nombre_de_hoja(titulo: str) -> str:
    """Nombre de hoja válido: sin caracteres prohibidos y de 31 caracteres o menos."""
    return (titulo or "").translate(_PROHIBIDOS_EN_HOJA).strip()[:31] or "Informe"


def construir_libro(titulo, columnas, filas, subtitulo=""):
    """Arma el libro de Excel completo y lo devuelve en memoria."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = nombre_de_hoja(titulo)

    _escribir_encabezado(hoja, titulo, subtitulo, len(columnas))

    fila_titulos = 4
    _escribir_titulos(hoja, columnas, fila_titulos)
    fila_siguiente = _escribir_filas(hoja, columnas, filas, fila_titulos + 1)

    if filas:
        _escribir_totales(hoja, columnas, filas, fila_siguiente)
        ultima_fila_datos = fila_siguiente - 1
        hoja.auto_filter.ref = f"A{fila_titulos}:{get_column_letter(len(columnas))}{ultima_fila_datos}"

    # Los títulos quedan a la vista al bajar por el listado
    hoja.freeze_panes = hoja.cell(row=fila_titulos + 1, column=1)
    hoja.sheet_view.showGridLines = False

    memoria = BytesIO()
    libro.save(memoria)
    memoria.seek(0)
    return memoria


def responder_excel(nombre_archivo, titulo, columnas, filas, subtitulo=""):
    """Respuesta HTTP con el informe listo para descargar."""
    memoria = construir_libro(titulo, columnas, filas, subtitulo)

    respuesta = make_response(memoria.read())
    respuesta.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    respuesta.headers["Content-Disposition"] = (
        f"attachment; filename={nombre_archivo}-{date.today().isoformat()}.xlsx"
    )
    return respuesta


def responder_plantilla_excel(nombre_archivo, hoja_titulo, encabezados, fila_ejemplo=None):
    """Plantilla simple para rellenar y volver a subir: encabezados en la fila 1, sin nada
    más encima. A diferencia de responder_excel(), no lleva el bloque de marca — los
    importadores del sitio leen la primera fila como encabezados, así que ponerle algo
    antes rompería la re-subida."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = hoja_titulo[:31] or "Plantilla"

    for indice, titulo in enumerate(encabezados, start=1):
        celda = hoja.cell(row=1, column=indice, value=titulo)
        celda.font = Font(bold=True, color="FFFFFF", size=10)
        celda.fill = PatternFill("solid", fgColor=_AZUL)
        hoja.column_dimensions[get_column_letter(indice)].width = max(14, len(titulo) + 2)

    if fila_ejemplo:
        for indice, valor in enumerate(fila_ejemplo, start=1):
            hoja.cell(row=2, column=indice, value=valor)

    hoja.freeze_panes = "A2"

    memoria = BytesIO()
    libro.save(memoria)
    memoria.seek(0)

    respuesta = make_response(memoria.read())
    respuesta.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    respuesta.headers["Content-Disposition"] = f"attachment; filename={nombre_archivo}.xlsx"
    return respuesta
