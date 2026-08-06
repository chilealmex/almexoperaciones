"""Lectura de la planilla "Provisión de Ingresos" (hoja Control).

La hoja trae una fila de totales arriba, los títulos en la fila 2 y los datos
desde la 3. Este módulo solo convierte esa hoja en diccionarios; quién los
guarda y cómo decide lo duplicado es cosa de las rutas.
"""

from datetime import date, datetime

HOJA = "Control"
# En la planilla de origen los títulos van en la fila 2, debajo de una fila de
# totales; en la plantilla que entrega el sistema van en la 1. Se busca la fila
# de títulos entre las primeras, así sirven las dos formas.
FILAS_TITULOS_POSIBLES = (1, 2, 3)

# Título en la planilla -> nombre del campo. Se busca por título y no por
# posición, así un cambio de orden de columnas no rompe la importación.
COLUMNAS = {
    "mes.año": "mes_ano",
    "mes.ano": "mes_ano",
    # El mes y el año también se aceptan en columnas separadas, que es como se
    # ven en pantalla y como los entrega la plantilla del sistema.
    "mes": "mes",
    "año": "anio",
    "ano": "anio",
    "cbte prov": "cbte_prov",
    "ot": "ot",
    "monto provisión": "monto_provision",
    "monto provision": "monto_provision",
    "reversa": "reversa",
    "mes reversa": "mes_reversa",
    "cbte reversa": "cbte_reversa",
    "cliente": "cliente",
    "centro de costos": "centro_costos",
    "rut": "rut",
    "obs": "obs",
    "saldo": "saldo",
}
# El período se puede indicar de dos formas: una sola columna con la fecha, o
# el mes y el año por separado. Basta con una de las dos.
OBLIGATORIAS = ("cbte_prov", "ot")
COLUMNAS_DE_PERIODO = ("mes_ano", "mes")

MESES_CORTOS = ("ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic")
MESES_NOMBRES = (
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
)


class PlanillaInvalida(Exception):
    """La planilla no tiene la forma esperada."""


def _texto(valor):
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def _entero(valor):
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return int(round(valor))
    texto = str(valor).replace("$", "").replace(".", "").replace(" ", "").replace(",", ".").strip()
    if not texto:
        return None
    try:
        return int(round(float(texto)))
    except ValueError:
        return None


def _fecha(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return None


def _periodo(fecha_completa, mes, anio):
    """El período de la línea, venga como fecha o como mes y año por separado."""
    fecha = _fecha(fecha_completa)
    if fecha:
        return fecha
    numero_mes = _entero(mes)
    numero_anio = _entero(anio)
    if not numero_mes or not numero_anio:
        return None
    if not (1 <= numero_mes <= 12):
        return None
    return date(numero_anio, numero_mes, 1)


def _mes_como_texto(valor):
    """El mes de reversa puede ser una fecha o texto libre ('may-26 y jun-26')."""
    fecha = _fecha(valor)
    if fecha:
        return f"{fecha.month:02d}.{fecha.year}"
    return _texto(valor)


def _columnas_de_la_fila(hoja, numero_fila):
    mapa = {}
    for celda in hoja[numero_fila]:
        titulo = _texto(celda.value)
        if not titulo:
            continue
        campo = COLUMNAS.get(titulo.lower())
        if campo and campo not in mapa:
            mapa[campo] = celda.column - 1
    return mapa


def _mapa_de_columnas(hoja):
    """Encuentra la fila de títulos y devuelve (mapa de columnas, número de fila)."""
    mejor, mejor_fila = {}, FILAS_TITULOS_POSIBLES[0]
    for numero_fila in FILAS_TITULOS_POSIBLES:
        if numero_fila > hoja.max_row:
            break
        mapa = _columnas_de_la_fila(hoja, numero_fila)
        if all(c in mapa for c in OBLIGATORIAS) and any(c in mapa for c in COLUMNAS_DE_PERIODO):
            return mapa, numero_fila
        if len(mapa) > len(mejor):
            mejor, mejor_fila = mapa, numero_fila

    nombres = {"cbte_prov": "Cbte Prov", "ot": "OT"}
    faltantes = [nombres[c] for c in OBLIGATORIAS if c not in mejor]
    if not any(c in mejor for c in COLUMNAS_DE_PERIODO):
        faltantes.append("Mes (o Mes.año)")
    raise PlanillaInvalida(
        "A la hoja 'Control' le faltan columnas obligatorias: " + ", ".join(faltantes)
    )


def leer_provisiones(archivo):
    """Devuelve una lista de diccionarios, uno por línea con datos de la hoja Control."""
    from openpyxl import load_workbook

    try:
        libro = load_workbook(archivo, data_only=True, read_only=True)
    except Exception as exc:  # archivo corrupto o que no es un xlsx
        raise PlanillaInvalida(f"No se pudo abrir el archivo: {exc}") from exc

    if HOJA not in libro.sheetnames:
        raise PlanillaInvalida(
            f"El archivo no tiene la hoja '{HOJA}'. Hojas encontradas: {', '.join(libro.sheetnames)}."
        )
    hoja = libro[HOJA]
    mapa, fila_titulos = _mapa_de_columnas(hoja)

    def valor(fila, campo):
        indice = mapa.get(campo)
        if indice is None or indice >= len(fila):
            return None
        return fila[indice]

    lineas = []
    descartadas = 0
    for numero, fila in enumerate(hoja.iter_rows(min_row=fila_titulos + 1, values_only=True), start=fila_titulos + 1):
        if not any(v not in (None, "") for v in fila):
            continue
        mes_ano = _periodo(valor(fila, "mes_ano"), valor(fila, "mes"), valor(fila, "anio"))
        cbte_prov = _texto(valor(fila, "cbte_prov"))
        ot = _texto(valor(fila, "ot"))
        if not (mes_ano and cbte_prov and ot):
            descartadas += 1
            continue  # fila de totales o incompleta: se ignora
        lineas.append(
            {
                "fila": numero,
                "mes_ano": mes_ano,
                "cbte_prov": cbte_prov,
                "ot": ot,
                "monto_provision": _entero(valor(fila, "monto_provision")) or 0,
                "reversa": _entero(valor(fila, "reversa")),
                "mes_reversa": _mes_como_texto(valor(fila, "mes_reversa")),
                "cbte_reversa": _texto(valor(fila, "cbte_reversa")),
                "cliente": _texto(valor(fila, "cliente")),
                "centro_costos": _texto(valor(fila, "centro_costos")),
                "rut": _texto(valor(fila, "rut")),
                "obs": _texto(valor(fila, "obs")),
                "saldo": _entero(valor(fila, "saldo")) or 0,
            }
        )
    libro.close()

    # Si había filas con datos y ninguna sirvió, casi siempre es porque los
    # títulos y los datos no están en la misma columna: pasa al insertar una
    # columna en la fila de títulos sin correr también los datos.
    if not lineas and descartadas:
        raise PlanillaInvalida(
            f"Se encontraron {descartadas} fila(s) con datos, pero ninguna tiene Mes, Cbte Prov y OT "
            "en las columnas que indican los títulos. Revisa que cada dato esté justo debajo de su "
            "título: si se insertó una columna, los títulos y los datos pueden haber quedado corridos."
        )
    return lineas


def mes_legible(fecha):
    """05.2026 -> 'may-26', para mostrar el mes como en la planilla."""
    if not fecha:
        return ""
    return f"{MESES_CORTOS[fecha.month - 1]}-{fecha.year % 100:02d}"


def nombre_mes(fecha):
    """Solo el mes, con nombre: 'Mayo'."""
    if not fecha:
        return ""
    return MESES_NOMBRES[fecha.month - 1]
