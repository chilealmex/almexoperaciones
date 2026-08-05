"""Lectura del mayor para el submódulo "Dif TC PR/CL".

La planilla trae los títulos en la primera fila y los datos desde la segunda.
Las columnas se buscan por título y no por posición, así un cambio de orden no
rompe la importación. El tipo de cambio vive en su propia columna (la celda de
la primera fila de datos) y se usa como valor inicial del período.
"""

from datetime import date, datetime

FILA_TITULOS = 1

COLUMNAS = {
    "cuenta": "cuenta",
    "descripción": "descripcion",
    "descripcion": "descripcion",
    "fecha": "fecha",
    "tipo": "tipo",
    "número": "numero",
    "numero": "numero",
    "id ficha": "id_ficha",
    "ficha": "ficha",
    "cargo ($)": "cargo",
    "abono ($)": "abono",
    "saldo ($)": "saldo",
    "código doc.": "codigo_doc",
    "codigo doc.": "codigo_doc",
    "documento": "documento",
    "vencimiento": "vencimiento",
    "número doc.": "numero_doc",
    "numero doc.": "numero_doc",
    "tipo mov.": "tipo_mov",
    "serie": "serie",
    "número mov.": "numero_mov",
    "numero mov.": "numero_mov",
    "moneda ref.": "moneda_ref",
    "comentario": "comentario",
    "doc. pago": "doc_pago",
    "número doc. pago": "numero_doc_pago",
    "numero doc. pago": "numero_doc_pago",
    "serie doc. pago.": "serie_doc_pago",
    "serie doc. pago": "serie_doc_pago",
    "mon orig": "mon_orig",
    "tipo de cambio": "tipo_cambio",
}
TEXTOS = (
    "cuenta", "descripcion", "tipo", "numero", "id_ficha", "ficha", "codigo_doc",
    "documento", "vencimiento", "numero_doc", "tipo_mov", "serie", "numero_mov",
    "moneda_ref", "comentario", "doc_pago", "numero_doc_pago", "serie_doc_pago",
)
ENTEROS = ("cargo", "abono", "saldo")
OBLIGATORIAS = ("cuenta", "saldo")


class PlanillaInvalida(Exception):
    """La planilla no tiene la forma esperada."""


def _texto(valor):
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def _numero(valor):
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).replace("$", "").replace(" ", "").strip()
    if not texto:
        return None
    # Un punto como separador de miles solo cuenta si además hay una coma decimal.
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return None


def _entero(valor):
    numero = _numero(valor)
    return int(round(numero)) if numero is not None else None


def _fecha(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return None


def _mapa_de_columnas(hoja):
    mapa = {}
    for celda in hoja[FILA_TITULOS]:
        titulo = _texto(celda.value)
        if not titulo:
            continue
        campo = COLUMNAS.get(titulo.lower())
        if campo and campo not in mapa:
            mapa[campo] = celda.column - 1
    faltantes = [c for c in OBLIGATORIAS if c not in mapa]
    if faltantes:
        raise PlanillaInvalida(
            "A la planilla le faltan columnas obligatorias: "
            + ", ".join(faltantes).replace("cuenta", "Cuenta").replace("saldo", "Saldo ($)")
        )
    return mapa


def leer_mayor(archivo, nombre_hoja=None):
    """Devuelve (líneas, tipo_de_cambio) leídos de la planilla."""
    from openpyxl import load_workbook

    try:
        libro = load_workbook(archivo, data_only=True, read_only=True)
    except Exception as exc:
        raise PlanillaInvalida(f"No se pudo abrir el archivo: {exc}") from exc

    if nombre_hoja and nombre_hoja in libro.sheetnames:
        hoja = libro[nombre_hoja]
    else:
        hoja = libro[libro.sheetnames[0]]
    mapa = _mapa_de_columnas(hoja)

    def valor(fila, campo):
        indice = mapa.get(campo)
        if indice is None or indice >= len(fila):
            return None
        return fila[indice]

    lineas = []
    tipo_cambio = None
    for numero_fila, fila in enumerate(
        hoja.iter_rows(min_row=FILA_TITULOS + 1, values_only=True), start=FILA_TITULOS + 1
    ):
        if not any(v not in (None, "") for v in fila):
            continue
        if tipo_cambio is None:
            tipo_cambio = _numero(valor(fila, "tipo_cambio"))
        if not _texto(valor(fila, "cuenta")):
            continue  # fila de totales o sin datos de mayor

        datos = {"orden": len(lineas)}
        for campo in TEXTOS:
            datos[campo] = _texto(valor(fila, campo))
        for campo in ENTEROS:
            datos[campo] = _entero(valor(fila, campo)) or 0
        datos["fecha"] = _fecha(valor(fila, "fecha"))
        datos["mon_orig"] = _numero(valor(fila, "mon_orig"))
        lineas.append(datos)

    libro.close()
    return lineas, tipo_cambio
