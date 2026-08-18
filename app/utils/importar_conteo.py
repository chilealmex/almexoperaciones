import csv
import io
import re
import unicodedata

from sqlalchemy import update

from app.extensions import db
from app.models.conteo_inventario import ItemConteoInventario
from app.utils.cantidades import a_cantidad

_ESPACIOS_RE = re.compile(r"\s+")
CERO = a_cantidad(0)


def _a_cantidad(valor):
    """Convierte un número de stock a Decimal, conservando los decimales.

    Acepta tanto texto (CSV) como números nativos (Excel vía openpyxl). Los
    decimales no se redondean: hay artículos que se miden en metros o kilos y
    "12,5" es el stock real, no un entero mal escrito.
    """
    return a_cantidad(valor, por_defecto=CERO)


def _a_monto(valor):
    """Convierte un importe exportado por planilla a entero de pesos.

    Acepta números nativos (Excel) y texto con distintos formatos de miles/decimales:
    '19,563,554', '1.961.923', '1.234,56' y '1,234.56' (el último separador manda
    cuando quedan 1-2 dígitos a su derecha; en cualquier otro caso son separadores de
    miles). Devuelve None si no hay dato utilizable.
    """
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return round(valor)

    limpio = str(valor).strip().replace("$", "").replace(" ", "").replace("\xa0", "")
    if not limpio:
        return None
    negativo = limpio.startswith("-") or (limpio.startswith("(") and limpio.endswith(")"))
    limpio = limpio.strip("-()")

    ultimo_separador = max(limpio.rfind(","), limpio.rfind("."))
    if ultimo_separador != -1 and len(limpio) - ultimo_separador - 1 in (1, 2):
        entero = limpio[:ultimo_separador].replace(".", "").replace(",", "")
        decimales = limpio[ultimo_separador + 1 :]
        numero_texto = f"{entero}.{decimales}" if entero else f"0.{decimales}"
    else:
        numero_texto = limpio.replace(".", "").replace(",", "")

    try:
        numero = round(float(numero_texto))
    except ValueError:
        return None
    return -numero if negativo else numero


def _buscar_columna(campos, *palabras_clave):
    """Primera columna cuyo nombre contiene todas las palabras clave (sin distinguir mayúsculas)."""
    for campo in campos or []:
        nombre = campo.lower()
        if all(palabra in nombre for palabra in palabras_clave):
            return campo
    return None


def _normalizar_codigo(codigo) -> str:
    """Limpia el código tal como se va a GUARDAR.

    Quita apóstrofes iniciales (artefacto de Excel), todos los espacios internos
    y los caracteres invisibles de categoría Cf (espacio de ancho cero, guion
    suave, BOM). QMS suele exportar 'ROP- BCAN-M' donde Defontana trae
    'ROP-BCAN-M' para el mismo código, y sin esta limpieza aparecen como dos
    artículos distintos en el cruce; con los invisibles es peor todavía, porque
    en pantalla se ven exactamente iguales.

    El guion NO se toca aquí: se deja el que venga en la planilla para que el
    código guardado siga siendo el del sistema de origen. Unificar los guiones
    es cosa de codigo_normalizado(), que es sólo la clave de comparación.
    """
    texto = str(codigo).strip().lstrip("'").strip()
    sin_invisibles = "".join(c for c in texto if unicodedata.category(c) != "Cf")
    return _ESPACIOS_RE.sub("", sin_invisibles)[:80]


def _texto(valor, limite: int) -> str:
    """Recorta al límite de la columna: PostgreSQL rechaza los textos que se pasan, SQLite no."""
    if valor is None:
        return ""
    return str(valor).strip()[:limite]


def _es_xlsx(file_storage) -> bool:
    nombre = (file_storage.filename or "").lower()
    if nombre.endswith(".xlsx"):
        return True
    if nombre.endswith(".csv"):
        return False
    # Sin extensión reconocible: un .xlsx es un zip, empieza con "PK".
    inicio = file_storage.stream.read(2)
    file_storage.stream.seek(0)
    return inicio == b"PK"


def _filas_desde_xlsx(file_storage):
    """(encabezados, filas) desde un .xlsx: cada fila es un dict encabezado -> valor nativo."""
    from openpyxl import load_workbook

    libro = load_workbook(file_storage.stream, data_only=True, read_only=True)
    hoja = libro.active
    filas = hoja.iter_rows(values_only=True)
    encabezados = [str(c).strip() if c is not None else "" for c in next(filas, [])]

    def generador():
        for fila in filas:
            if all(valor is None for valor in fila):
                continue
            yield dict(zip(encabezados, fila))

    return encabezados, generador()


def _filas_desde_csv(file_storage, codificaciones=("utf-8-sig",)):
    """(encabezados, filas) desde un .csv separado por punto y coma, con las codificaciones dadas en orden."""
    crudo = file_storage.stream.read()
    for codificacion in codificaciones:
        try:
            contenido = crudo.decode(codificacion)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("No se pudo leer la codificación del archivo.")

    reader = csv.DictReader(io.StringIO(contenido), delimiter=";")
    return reader.fieldnames, reader


def _leer_filas(file_storage, codificaciones_csv=("utf-8-sig",)):
    """Punto único de entrada: detecta si el archivo es .xlsx o .csv y devuelve (encabezados, filas)."""
    if _es_xlsx(file_storage):
        return _filas_desde_xlsx(file_storage)
    return _filas_desde_csv(file_storage, codificaciones_csv)


# IMPORTANTE: los importadores sólo actualizan lo que declara cada sistema
# (cantidades, costo, unidad, nombre, ubicación). Nunca escriben cantidad_fisica,
# contado_por_id ni contado_en: el conteo físico es trabajo de bodega y debe
# sobrevivir a cualquier reimportación de QMS o Defontana.


# Guiones que Excel y Word escriben en vez del normal. Se unifican al '-' de
# toda la vida: "KIT–ST" y "KIT-ST" son el mismo artículo escrito distinto.
_GUIONES = {
    "‐": "-",  # HYPHEN
    "‑": "-",  # NON-BREAKING HYPHEN
    "‒": "-",  # FIGURE DASH
    "–": "-",  # EN DASH
    "—": "-",  # EM DASH
    "―": "-",  # HORIZONTAL BAR
    "−": "-",  # MINUS SIGN
}


def codigo_normalizado(codigo) -> str:
    """Clave para comparar códigos que son el mismo escrito distinto.

    QMS y Defontana no siempre escriben igual el código del mismo artículo:
    "EM-R-Pantalla BG3" y "EM-R-PantallaBG3" son el mismo producto. Para
    compararlos se saca todo lo que no cambia de qué artículo se trata:

    - espacios de cualquier tipo, incluido el espacio duro que pega Excel;
    - caracteres invisibles (categoría Cf): espacio de ancho cero, guion suave,
      la marca BOM... No se ven en pantalla, así que dos códigos que sólo se
      diferencian en eso parecen idénticos y aun así no cruzaban;
    - los distintos guiones tipográficos, que se unifican al '-';
    - el apóstrofe con que Excel marca "esto es texto";
    - acentos y mayúsculas.
    """
    texto = str(codigo or "").strip().lstrip("'").strip()
    # Los Cf hay que sacarlos antes de comparar nada: son invisibles.
    sin_invisibles = "".join(c for c in texto if unicodedata.category(c) != "Cf")
    sin_guiones = "".join(_GUIONES.get(c, c) for c in sin_invisibles)
    sin_acentos = "".join(
        c for c in unicodedata.normalize("NFD", sin_guiones) if unicodedata.category(c) != "Mn"
    )
    return "".join(sin_acentos.split()).upper()


def _congela_el_stock(item, solo_no_contados: bool) -> bool:
    """True si a este artículo no hay que actualizarle el stock declarado por el sistema.

    Cuando la toma de inventario dura varios días, bodega cuenta un artículo el
    lunes contra el stock que el sistema declaraba el lunes. Si el martes se
    reimporta el stock, ese artículo pasaría a compararse contra una cifra que ya
    incorpora los movimientos del martes, y aparecería una diferencia que no
    existe: la toma dejaría de cuadrar por culpa de la reimportación, no por un
    error de bodega.

    Por eso, con solo_no_contados=True, a los artículos ya contados se les
    congela la cantidad: siguen comparándose contra la foto del sistema que
    tenían al momento de contarlos. Los que aún no se cuentan sí se actualizan,
    que es justamente lo que se quiere refrescar antes de seguir contando.

    Ojo: sólo se congela la CANTIDAD. Nombre, ubicación, unidad y costo se
    siguen actualizando, porque son datos de referencia que no cambian el
    resultado del conteo (no son movimientos de stock).
    """
    return solo_no_contados and item.cantidad_fisica is not None


def _aplicar_en_lote(cambios: list) -> None:
    """Aplica muchas actualizaciones en una sola ida y vuelta a la base.

    Modificar los objetos del ORM uno por uno parece equivalente, pero al
    guardar SQLAlchemy agrupa los UPDATE por el conjunto de columnas que cambió
    en cada fila. Como aquí cada artículo cambia columnas distintas (el nombre
    sólo si viene, el costo sólo si viene...), se fragmentaba en más de mil
    sentencias, y contra la base remota de Render cada una es una ida y vuelta
    por la red: de ahí que importar tardara minutos.

    Mandando siempre las mismas claves en todos los diccionarios, SQLAlchemy
    puede usar executemany y resolverlo con una sentencia.
    """
    if not cambios:
        return
    # En tandas para no armar una sentencia gigante con miles de parámetros.
    for inicio in range(0, len(cambios), 500):
        db.session.execute(update(ItemConteoInventario), cambios[inicio : inicio + 500])


def _marcar_ausentes_en_lote(items, codigos_del_archivo, campo: str) -> None:
    """Apaga la marca del sistema recién importado en los artículos que no venían.

    'items' son los que ya existían antes de esta importación, así que los
    recién creados quedan fuera por construcción: un artículo creado por ESTA
    planilla no puede estar ausente de ella. El otro sistema no se toca, porque
    de esa planilla no sabemos nada en esta pasada.
    """
    presentes = {codigo_normalizado(c) for c in codigos_del_archivo}
    ids = [
        item.id
        for item in items
        if getattr(item, campo) and codigo_normalizado(item.codigo) not in presentes
    ]
    for inicio in range(0, len(ids), 500):
        db.session.execute(
            update(ItemConteoInventario)
            .where(ItemConteoInventario.id.in_(ids[inicio : inicio + 500]))
            .values(**{campo: False})
        )


def _items_de_la_empresa(empresa_id: int) -> list:
    """Todos los artículos de la empresa, incluidos los códigos repetidos.

    Una sola consulta: pedirlos de a uno genera miles de idas y vueltas contra
    la base remota y agota el tiempo de la petición.
    """
    return ItemConteoInventario.query.filter_by(empresa_id=empresa_id).all()


def _por_codigo_normalizado(items) -> dict:
    """Indexa por código normalizado para cruzar QMS y Defontana como el mismo artículo.

    Si hay duplicados de antes, gana el que tenga conteo físico registrado (y
    entre iguales, el más antiguo), para no dejar huérfano el trabajo de bodega.
    """
    por_codigo = {}
    for item in sorted(items, key=lambda i: (i.cantidad_fisica is None, i.id)):
        por_codigo.setdefault(codigo_normalizado(item.codigo), item)
    return por_codigo


def importar_qms(file_storage, empresa_id: int, solo_no_contados: bool = False) -> dict:
    """Archivo 'Distribución Valor Stock CLP' de QMS (.csv o .xlsx): código único, descripción,
    stock, línea de negocio, categoría, unidad y costo unitario.

    Con solo_no_contados=True no se toca el stock de los artículos que ya tienen
    conteo físico registrado (ver _congela_el_stock).
    """
    encabezados, reader = _leer_filas(file_storage)

    columna_codigo = next((c for c in encabezados if "digo" in c.lower() and "nico" in c.lower()), None)
    columna_nombre = next((c for c in encabezados if "descripci" in c.lower()), None)
    columna_stock = "Stock" if "Stock" in encabezados else _buscar_columna(encabezados, "stock")
    columna_linea = "Linea Negocio" if "Linea Negocio" in encabezados else _buscar_columna(encabezados, "linea")
    columna_ubicacion = "ubicacion_bodega"
    columna_unidad = _buscar_columna(encabezados, "unidad")
    columna_categoria = _buscar_columna(encabezados, "categor")
    columna_costo = _buscar_columna(encabezados, "valor", "unitario")
    columna_valor_total = _buscar_columna(encabezados, "valor", "total")

    if not columna_codigo or not columna_nombre:
        raise ValueError("No se encontraron las columnas de código/descripción esperadas en el archivo QMS.")

    acumulado = {}  # codigo -> dict con datos agregados
    for fila in reader:
        codigo = _normalizar_codigo(fila.get(columna_codigo) or "")
        if not codigo:
            continue
        cantidad = _a_cantidad(fila.get(columna_stock, "0"))
        if codigo not in acumulado:
            acumulado[codigo] = {
                "cantidad": 0,
                "nombre": _texto(fila.get(columna_nombre), 255),
                "linea_negocio": _texto(fila.get(columna_linea), 120) if columna_linea else "",
                "ubicacion": _texto(fila.get(columna_ubicacion), 255) or _texto(fila.get("Sucursal"), 255),
                "categoria": _texto(fila.get(columna_categoria), 120) if columna_categoria else "",
                "unidad": _texto(fila.get(columna_unidad), 20) if columna_unidad else "",
                "costo": None,
            }
        acumulado[codigo]["cantidad"] += cantidad

        # Costo unitario: viene explícito o se deduce del valor total de la fila
        costo = _a_monto(fila.get(columna_costo)) if columna_costo else None
        if costo is None and columna_valor_total and cantidad:
            total = _a_monto(fila.get(columna_valor_total))
            if total:
                costo = round(total / cantidad)
        if costo and acumulado[codigo]["costo"] is None:
            acumulado[codigo]["costo"] = costo

    items = _items_de_la_empresa(empresa_id)
    existentes = _por_codigo_normalizado(items)
    filas_creadas = 0
    filas_actualizadas = 0
    filas_congeladas = 0
    nuevos = []
    cambios = []
    for codigo, datos in acumulado.items():
        item = existentes.get(codigo_normalizado(codigo))
        if item is None:
            item = ItemConteoInventario(empresa_id=empresa_id, codigo=codigo, cantidad_defontana=0,
                                        en_qms=True, en_defontana=False)
            item.cantidad_qms = datos["cantidad"]
            item.nombre = datos["nombre"] or None
            item.linea_negocio = datos["linea_negocio"] or None
            item.ubicacion = datos["ubicacion"] or None
            item.categoria = datos["categoria"] or None
            item.unidad_qms = datos["unidad"] or None
            item.costo_unitario_qms = datos["costo"]
            nuevos.append(item)
            filas_creadas += 1
            continue

        congelado = _congela_el_stock(item, solo_no_contados)
        if congelado:
            filas_congeladas += 1
        else:
            filas_actualizadas += 1
        # Se manda siempre el mismo juego de columnas (con el valor que ya tenía
        # cuando la planilla no trae dato) para que salga en una sola sentencia.
        cambios.append({
            "id": item.id,
            "cantidad_qms": item.cantidad_qms if congelado else datos["cantidad"],
            "nombre": datos["nombre"] or item.nombre,
            "linea_negocio": datos["linea_negocio"] or item.linea_negocio,
            "ubicacion": item.ubicacion or datos["ubicacion"] or item.ubicacion,
            "categoria": datos["categoria"] or item.categoria,
            "unidad_qms": datos["unidad"] or item.unidad_qms,
            "costo_unitario_qms": datos["costo"] if datos["costo"] is not None else item.costo_unitario_qms,
            "en_qms": True,
        })

    if nuevos:
        db.session.add_all(nuevos)
    _aplicar_en_lote(cambios)
    _marcar_ausentes_en_lote(items, acumulado.keys(), "en_qms")
    db.session.commit()
    return {
        "total_codigos": len(acumulado),
        "creados": filas_creadas,
        "actualizados": filas_actualizadas,
        "congelados": filas_congeladas,
    }


def importar_defontana(file_storage, empresa_id: int, solo_no_contados: bool = False) -> dict:
    """Archivo de inventario por bodega de Defontana (.csv o .xlsx): CodArticulo, Descripción,
    CodBodega, Nombre Bodega, Saldo Stock."""
    encabezados, reader = _leer_filas(file_storage, codificaciones_csv=("cp1252", "latin-1", "utf-8-sig"))

    columna_codigo = next((c for c in encabezados if "codarticulo" in c.lower()), None)
    columna_nombre = next((c for c in encabezados if "descripci" in c.lower()), None)
    columna_stock = next((c for c in encabezados if "saldo" in c.lower()), None)
    columna_bodega = next((c for c in encabezados if "nombre bodega" in c.lower()), None)
    columna_unidad = _buscar_columna(encabezados, "unidad")
    # Defontana exporta el costo con nombres distintos según el informe
    columna_costo = (
        _buscar_columna(encabezados, "costo", "unitario")
        or _buscar_columna(encabezados, "costo", "promedio")
        or _buscar_columna(encabezados, "costo")
        or _buscar_columna(encabezados, "valor", "unitario")
    )
    columna_valor_total = _buscar_columna(encabezados, "valor", "total") or _buscar_columna(
        encabezados, "total", "costo"
    )

    if not columna_codigo or not columna_stock:
        raise ValueError("No se encontraron las columnas de código/stock esperadas en el archivo de Defontana.")

    acumulado = {}
    for fila in reader:
        codigo = _normalizar_codigo(fila.get(columna_codigo) or "")
        if not codigo:
            continue
        cantidad = _a_cantidad(fila.get(columna_stock, "0"))
        bodega = _texto(fila.get(columna_bodega), 255) if columna_bodega else ""
        if codigo not in acumulado:
            acumulado[codigo] = {
                "cantidad": 0,
                "nombre": _texto(fila.get(columna_nombre), 255) if columna_nombre else "",
                "bodegas": set(),
                "unidad": _texto(fila.get(columna_unidad), 20) if columna_unidad else "",
                "costo": None,
            }
        acumulado[codigo]["cantidad"] += cantidad
        if bodega and cantidad:
            acumulado[codigo]["bodegas"].add(bodega)

        costo = _a_monto(fila.get(columna_costo)) if columna_costo else None
        if costo is None and columna_valor_total and cantidad:
            total = _a_monto(fila.get(columna_valor_total))
            if total:
                costo = round(total / cantidad)
        if costo and acumulado[codigo]["costo"] is None:
            acumulado[codigo]["costo"] = costo

    items = _items_de_la_empresa(empresa_id)
    existentes = _por_codigo_normalizado(items)
    filas_creadas = 0
    filas_actualizadas = 0
    filas_congeladas = 0
    nuevos = []
    cambios = []
    for codigo, datos in acumulado.items():
        bodegas = ", ".join(sorted(datos["bodegas"]))[:255] if datos["bodegas"] else ""
        item = existentes.get(codigo_normalizado(codigo))
        if item is None:
            item = ItemConteoInventario(empresa_id=empresa_id, codigo=codigo, cantidad_qms=0,
                                        en_defontana=True, en_qms=False)
            item.cantidad_defontana = datos["cantidad"]
            item.nombre = datos["nombre"] or None
            item.ubicacion = bodegas or None
            item.unidad_defontana = datos["unidad"] or None
            item.costo_unitario_defontana = datos["costo"]
            nuevos.append(item)
            filas_creadas += 1
            continue

        congelado = _congela_el_stock(item, solo_no_contados)
        if congelado:
            filas_congeladas += 1
        else:
            filas_actualizadas += 1
        cambios.append({
            "id": item.id,
            "cantidad_defontana": item.cantidad_defontana if congelado else datos["cantidad"],
            "nombre": item.nombre or datos["nombre"] or item.nombre,
            "ubicacion": item.ubicacion or bodegas or item.ubicacion,
            "unidad_defontana": datos["unidad"] or item.unidad_defontana,
            "costo_unitario_defontana": (
                datos["costo"] if datos["costo"] is not None else item.costo_unitario_defontana
            ),
            "en_defontana": True,
        })

    if nuevos:
        db.session.add_all(nuevos)
    _aplicar_en_lote(cambios)
    _marcar_ausentes_en_lote(items, acumulado.keys(), "en_defontana")
    db.session.commit()
    return {
        "total_codigos": len(acumulado),
        "creados": filas_creadas,
        "actualizados": filas_actualizadas,
        "congelados": filas_congeladas,
    }


def articulos_fuera_de_ambas_planillas(empresa_id: int) -> list:
    """Artículos que dejaron de aparecer tanto en QMS como en Defontana."""
    return (
        ItemConteoInventario.query.filter_by(empresa_id=empresa_id, en_qms=False, en_defontana=False)
        .order_by(ItemConteoInventario.codigo)
        .all()
    )


def grupos_duplicados(empresa_id: int) -> list:
    """Artículos cuyo código es el mismo si se le quitan espacios y acentos.

    Devuelve una lista de grupos (2 o más artículos cada uno), ordenados por
    código, para poder revisarlos y unificarlos desde la pantalla.
    """
    from collections import defaultdict as _dd

    por_clave = _dd(list)
    for item in ItemConteoInventario.query.filter_by(empresa_id=empresa_id).all():
        por_clave[codigo_normalizado(item.codigo)].append(item)
    grupos = [sorted(items, key=lambda i: i.id) for items in por_clave.values() if len(items) > 1]
    return sorted(grupos, key=lambda g: g[0].codigo)


def _primero(valores):
    return next((v for v in valores if v not in (None, "")), None)


def unificar_grupo(items: list) -> ItemConteoInventario:
    """Deja un solo artículo con la suma de los duplicados y borra el resto.

    Se conserva la fila que ya tenía conteo físico (y entre iguales, la más
    antigua), para no perder el trabajo de bodega. Las cantidades se suman,
    porque cada fila era stock declarado por separado, y los datos de texto que
    falten en la fila que queda se completan con los de las otras.
    """
    if len(items) < 2:
        return items[0] if items else None

    ordenados = sorted(items, key=lambda i: (i.cantidad_fisica is None, i.id))
    principal, resto = ordenados[0], ordenados[1:]

    principal.cantidad_qms = sum(i.cantidad_qms or 0 for i in ordenados)
    principal.cantidad_defontana = sum(i.cantidad_defontana or 0 for i in ordenados)

    contadas = [i.cantidad_fisica for i in ordenados if i.cantidad_fisica is not None]
    principal.cantidad_fisica = sum(contadas) if contadas else None
    contador = next((i for i in ordenados if i.cantidad_fisica is not None), None)
    if contador is not None:
        principal.contado_por_id = contador.contado_por_id
        principal.contado_en = contador.contado_en

    for campo in ("nombre", "linea_negocio", "ubicacion", "categoria", "unidad_qms",
                  "unidad_defontana", "costo_unitario_qms", "costo_unitario_defontana"):
        if getattr(principal, campo) in (None, ""):
            setattr(principal, campo, _primero(getattr(i, campo) for i in ordenados))

    for item in resto:
        db.session.delete(item)
    return principal


# Nombre corto para los caracteres que no se ven pero separan dos códigos.
_INVISIBLES = {
    " ": "espacio duro",
    "​": "espacio de ancho cero",
    "‌": "separador de ancho cero",
    "‍": "unión de ancho cero",
    "⁠": "unión invisible",
    "﻿": "marca BOM",
    "­": "guion suave",
    "\t": "tabulador",
}


def rarezas_del_codigo(codigo) -> list:
    """Qué tiene este código que no se ve en pantalla.

    Dos códigos que sólo se diferencian en un carácter invisible se ven
    idénticos, así que en la pantalla de depuración no habría forma de saber
    por qué aparecen repetidos. Esto lo explica en palabras.
    """
    texto = str(codigo or "")
    encontradas = []
    if texto != texto.strip():
        encontradas.append("espacios al principio o al final")
    if texto.startswith("'"):
        encontradas.append("apóstrofe de Excel")
    if " " in texto.strip():
        encontradas.append("espacios en medio")
    for caracter, nombre in _INVISIBLES.items():
        if caracter in texto:
            encontradas.append(nombre)
    for caracter in _GUIONES:
        if caracter in texto:
            encontradas.append("guion tipográfico")
            break
    return encontradas
