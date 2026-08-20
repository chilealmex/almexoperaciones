"""La pantalla de Conciliación SII / Defontana, de punta a punta.

Cubre lo que la usuaria pidió ver: que se carguen los cuatro archivos, que el
listado muestre los montos, que el detalle diga en qué se diferencia cada
documento y que la tabla traiga los totales de cada columna.
"""

import io

from werkzeug.datastructures import FileStorage

from app.models.contabilidad import ConciliacionSii, ConciliacionSiiDocumento
from tests.conftest import login

CSV_COMPRA = (
    "Nro;Tipo Doc;RUT Proveedor;Razon Social;Folio;Fecha Docto;"
    "Monto Exento;Monto Neto;Monto IVA Recuperable;Monto Total\r\n"
    "1;33;76.123.456-7;PROVEEDOR UNO SPA;1001;05/08/2026;0;100000;19000;119000\r\n"
    "2;33;77.222.333-4;PROVEEDOR DOS LTDA;2002;12/08/2026;0;50000;9500;59500\r\n"
)

CSV_VENTA = (
    "Nro;Tipo Doc;Rut cliente;Razon Social;Folio;Fecha Docto;"
    "Monto Exento;Monto Neto;Monto IVA;Monto total\r\n"
    "1;33;60.111.222-3;CLIENTE UNO SA;5001;03/08/2026;0;200000;38000;238000\r\n"
)

LIBRO_COMPRAS = """<html><body><table>
  <tr><td>Documento: 33</td></tr>
  <tr><td>Folio</td><td>Fecha</td><td>RUT</td><td>Razon Social</td>
      <td>Neto</td><td>Exento</td><td>IVA</td><td>Total</td></tr>
  <tr><td>1001</td><td>05/08/2026</td><td>76.123.456-7</td><td>PROVEEDOR UNO SPA</td>
      <td>90.000</td><td>0</td><td>17.100</td><td>107.100</td></tr>
</table></body></html>"""

LIBRO_VENTAS = """<html><body><table>
  <tr><td>Documento: 33</td></tr>
  <tr><td>Folio</td><td>Fecha</td><td>RUT</td><td>Razon Social</td>
      <td>Neto</td><td>Exento</td><td>IVA</td><td>Total</td></tr>
  <tr><td>5001</td><td>03/08/2026</td><td>60.111.222-3</td><td>CLIENTE UNO SA</td>
      <td>200.000</td><td>0</td><td>38.000</td><td>238.000</td></tr>
</table></body></html>"""


def _archivo(texto, nombre, codificacion="utf-8"):
    return FileStorage(stream=io.BytesIO(texto.encode(codificacion)), filename=nombre)


def _cargar(client, **extra):
    datos = {
        "anio": "2026",
        "mes": "8",
        "sii_compra": _archivo(CSV_COMPRA, "compra.csv"),
        "defontana_compra": _archivo(LIBRO_COMPRAS, "compras.xls", "cp1252"),
        "sii_venta": _archivo(CSV_VENTA, "venta.csv"),
        "defontana_venta": _archivo(LIBRO_VENTAS, "ventas.xls", "cp1252"),
    }
    datos.update(extra)
    return client.post(
        "/contabilidad/conciliacion-sii/cargar",
        data=datos, content_type="multipart/form-data", follow_redirects=True,
    )


def test_cargar_un_periodo_completo(client, db, empresa, usuario_admin):
    login(client, "admin@test.cl")

    respuesta = _cargar(client)

    assert respuesta.status_code == 200
    conciliacion = ConciliacionSii.query.one()
    assert (conciliacion.anio, conciliacion.mes) == (2026, 8)
    assert conciliacion.compras is not None
    assert conciliacion.ventas is not None

    compras = conciliacion.compras
    # 1001 está en los dos por montos distintos, 2002 sólo en el SII
    assert compras.n_dif_monto == 1
    assert compras.n_solo_sii == 1
    assert compras.total_sii == 178500      # 119.000 + 59.500
    assert compras.total_defontana == 107100
    assert compras.diferencia == 71400
    assert not compras.cuadra

    assert conciliacion.ventas.cuadra
    assert conciliacion.ventas.n_coincide == 1


def test_el_listado_muestra_los_montos(client, db, empresa, usuario_admin):
    """Lo que pidió la usuaria: ver el monto, no sólo el conteo de documentos."""
    login(client, "admin@test.cl")
    _cargar(client)

    body = client.get("/contabilidad/conciliacion-sii").get_data(as_text=True)

    assert "$178.500" in body      # total SII de compras
    assert "$107.100" in body      # total Defontana
    assert "$71.400" in body       # la diferencia
    assert "Agosto 2026" in body


def test_el_detalle_dice_en_que_se_diferencia(client, db, empresa, usuario_admin):
    login(client, "admin@test.cl")
    _cargar(client)
    conciliacion = ConciliacionSii.query.one()

    body = client.get(
        f"/contabilidad/conciliacion-sii/{conciliacion.id}/compra"
    ).get_data(as_text=True)

    assert "Diferencia de monto" in body
    assert "Neto: $100.000 vs $90.000" in body
    assert "IVA: $19.000 vs $17.100" in body
    # Y el que falta en Defontana muestra el monto que falta
    assert "Solo en SII" in body


def test_el_detalle_se_abre_con_un_boton_y_no_alarga_las_filas(client, db, empresa, usuario_admin):
    """El texto largo iba dentro de la fila y la hacía crecer a tres renglones."""
    login(client, "admin@test.cl")
    _cargar(client)
    conciliacion = ConciliacionSii.query.one()

    body = client.get(
        f"/contabilidad/conciliacion-sii/{conciliacion.id}/compra"
    ).get_data(as_text=True)

    # El detalle viaja en el botón, no suelto en la celda
    assert 'data-bs-target="#modal-diferencia"' in body
    assert 'data-detalle="Neto: $100.000 vs $90.000' in body
    # Y las razones sociales se recortan a una línea
    assert "celda-recortada" in body


def test_la_tabla_trae_los_totales_de_cada_columna(client, db, empresa, usuario_admin):
    """El otro pedido: los valores totales por columna."""
    login(client, "admin@test.cl")
    _cargar(client)
    conciliacion = ConciliacionSii.query.one()

    body = client.get(
        f"/contabilidad/conciliacion-sii/{conciliacion.id}/compra"
    ).get_data(as_text=True)

    assert "Totales (2 documentos)" in body
    for etiqueta in ("Neto SII", "Neto Defontana", "Dif. neto", "Exento SII",
                     "IVA SII", "IVA Defontana", "Dif. IVA", "Total SII", "Dif. total"):
        assert etiqueta in body, f"falta la columna {etiqueta}"


def test_los_totales_siguen_al_filtro(client, db, empresa, usuario_admin):
    login(client, "admin@test.cl")
    _cargar(client)
    conciliacion = ConciliacionSii.query.one()
    url = f"/contabilidad/conciliacion-sii/{conciliacion.id}/compra"

    completo = client.get(url).get_data(as_text=True)
    filtrado = client.get(f"{url}?filtro=solo_sii").get_data(as_text=True)

    assert "Totales (2 documentos)" in completo
    assert "Totales (1 documento)" in filtrado
    # Con el filtro puesto sólo queda el documento de $59.500
    assert "$178.500" in completo
    assert "$178.500" not in filtrado


def test_se_puede_cargar_solo_un_libro_y_el_otro_despues(client, db, empresa, usuario_admin):
    login(client, "admin@test.cl")

    _cargar(client, sii_venta=(io.BytesIO(b""), ""), defontana_venta=(io.BytesIO(b""), ""))
    conciliacion = ConciliacionSii.query.one()
    assert conciliacion.compras is not None
    assert conciliacion.ventas is None

    # El mismo mes, ahora con ventas: se suma al período que ya existía
    _cargar(client, sii_compra=(io.BytesIO(b""), ""), defontana_compra=(io.BytesIO(b""), ""))
    db.session.expire_all()
    assert ConciliacionSii.query.count() == 1
    assert ConciliacionSii.query.one().ventas is not None


def test_recargar_un_libro_reemplaza_el_cruce_anterior(client, db, empresa, usuario_admin):
    login(client, "admin@test.cl")
    _cargar(client)
    documentos_antes = ConciliacionSiiDocumento.query.count()

    _cargar(client)

    db.session.expire_all()
    assert ConciliacionSiiDocumento.query.count() == documentos_antes
    assert ConciliacionSii.query.one().compras.cargas == 2


def test_un_archivo_que_no_corresponde_avisa_y_no_guarda_nada(client, db, empresa, usuario_admin):
    login(client, "admin@test.cl")

    respuesta = _cargar(
        client,
        sii_compra=_archivo("a;b;c\r\n1;2;3\r\n", "cualquiera.csv"),
        defontana_compra=_archivo("<html><body></body></html>", "otro.xls", "cp1252"),
        sii_venta=(io.BytesIO(b""), ""),
        defontana_venta=(io.BytesIO(b""), ""),
    )

    assert respuesta.status_code == 200
    assert "Tipo Doc" in respuesta.get_data(as_text=True)
    assert ConciliacionSii.query.count() == 0


def test_falta_uno_de_los_dos_archivos_de_un_libro(client, db, empresa, usuario_admin):
    login(client, "admin@test.cl")

    respuesta = _cargar(
        client,
        defontana_compra=(io.BytesIO(b""), ""),
        sii_venta=(io.BytesIO(b""), ""),
        defontana_venta=(io.BytesIO(b""), ""),
    )

    assert "falta el libro de Defontana" in respuesta.get_data(as_text=True)
    assert ConciliacionSii.query.count() == 0


def test_el_excel_sale_con_los_totales(client, db, empresa, usuario_admin):
    from openpyxl import load_workbook

    login(client, "admin@test.cl")
    _cargar(client)
    conciliacion = ConciliacionSii.query.one()

    respuesta = client.get(
        f"/contabilidad/conciliacion-sii/{conciliacion.id}/compra.xlsx"
    )
    assert respuesta.status_code == 200

    hoja = load_workbook(io.BytesIO(respuesta.data)).active
    titulos = [c.value for c in hoja[4]]
    assert "Neto SII" in titulos and "Dif. total" in titulos
    assert "En qué se diferencia" in titulos

    columna_total_sii = titulos.index("Total SII") + 1
    assert hoja.cell(row=hoja.max_row, column=columna_total_sii).value == 178500


def test_eliminar_borra_el_periodo_completo(client, db, empresa, usuario_admin):
    login(client, "admin@test.cl")
    _cargar(client)
    conciliacion = ConciliacionSii.query.one()

    client.post(
        f"/contabilidad/conciliacion-sii/{conciliacion.id}/eliminar",
        data={}, follow_redirects=True,
    )

    assert ConciliacionSii.query.count() == 0
    assert ConciliacionSiiDocumento.query.count() == 0


def test_sin_permiso_de_contabilidad_no_se_entra(client, db, empresa, usuario_bodega):
    login(client, "bodega@test.cl")
    assert client.get("/contabilidad/conciliacion-sii").status_code == 403


def test_aparece_en_el_menu_de_contabilidad(client, db, empresa, usuario_admin):
    login(client, "admin@test.cl")
    body = client.get("/contabilidad/conciliacion-sii").get_data(as_text=True)
    assert "Conciliación SII" in body


# --- Aceptar una diferencia que es correcta ---
#
# Hay descuadres que no son errores: las facturas de combustible llevan
# impuesto específico y el SII y Defontana no lo reparten igual entre neto e
# impuestos. Sin poder darlos por buenos, el mes nunca llega a cero pendientes.


def _documento_con_diferencia(libro):
    return next(d for d in libro.documentos if d.estado == "dif_monto")


def _aceptar(client, conciliacion, documento, motivo=None):
    datos = {"motivo": motivo} if motivo is not None else {}
    return client.post(
        f"/contabilidad/conciliacion-sii/{conciliacion.id}/compra/{documento.id}/aceptar",
        data=datos, follow_redirects=True,
    )


def test_aceptar_una_diferencia_la_saca_de_pendientes(client, db, empresa, usuario_admin):
    login(client, "admin@test.cl")
    _cargar(client)
    conciliacion = ConciliacionSii.query.one()
    libro = conciliacion.compras
    pendientes_antes = libro.pendientes
    documento = _documento_con_diferencia(libro)

    _aceptar(client, conciliacion, documento)

    db.session.expire_all()
    conciliacion = ConciliacionSii.query.one()
    documento = db.session.get(ConciliacionSiiDocumento, documento.id)
    assert documento.aceptado is True
    assert documento.motivo_aceptacion == "Impuesto específico a los combustibles"
    assert documento.aceptado_por_id == usuario_admin.id
    assert documento.aceptado_en is not None
    assert conciliacion.compras.n_aceptados == 1
    assert conciliacion.compras.pendientes == pendientes_antes - 1


def test_aceptar_no_cambia_ningun_monto(client, db, empresa, usuario_admin):
    """Aceptar es dejar constancia de una revisión, no corregir la contabilidad."""
    login(client, "admin@test.cl")
    _cargar(client)
    conciliacion = ConciliacionSii.query.one()
    documento = _documento_con_diferencia(conciliacion.compras)
    antes = (documento.neto_sii, documento.neto_defontana, documento.total_sii,
             documento.total_defontana, documento.diferencia, documento.estado)

    _aceptar(client, conciliacion, documento)

    db.session.expire_all()
    documento = db.session.get(ConciliacionSiiDocumento, documento.id)
    despues = (documento.neto_sii, documento.neto_defontana, documento.total_sii,
               documento.total_defontana, documento.diferencia, documento.estado)
    assert antes == despues


def test_se_puede_escribir_otro_motivo(client, db, empresa, usuario_admin):
    login(client, "admin@test.cl")
    _cargar(client)
    conciliacion = ConciliacionSii.query.one()
    documento = _documento_con_diferencia(conciliacion.compras)

    _aceptar(client, conciliacion, documento, motivo="Nota de crédito emitida en agosto")

    db.session.expire_all()
    assert db.session.get(
        ConciliacionSiiDocumento, documento.id
    ).motivo_aceptacion == "Nota de crédito emitida en agosto"


def test_volver_a_pulsar_deshace_la_aceptacion(client, db, empresa, usuario_admin):
    login(client, "admin@test.cl")
    _cargar(client)
    conciliacion = ConciliacionSii.query.one()
    documento = _documento_con_diferencia(conciliacion.compras)
    _aceptar(client, conciliacion, documento)

    _aceptar(client, conciliacion, documento)

    db.session.expire_all()
    documento = db.session.get(ConciliacionSiiDocumento, documento.id)
    assert documento.aceptado is False
    assert documento.motivo_aceptacion is None
    assert ConciliacionSii.query.one().compras.n_aceptados == 0


def test_la_aceptacion_sobrevive_a_recargar_los_archivos(client, db, empresa, usuario_admin):
    """La conciliación se recarga varias veces en el mes; el trabajo no se pierde."""
    login(client, "admin@test.cl")
    _cargar(client)
    conciliacion = ConciliacionSii.query.one()
    documento = _documento_con_diferencia(conciliacion.compras)
    llave = (documento.tipo_doc, documento.folio)
    _aceptar(client, conciliacion, documento, motivo="Impuesto específico")

    _cargar(client)  # se vuelven a subir los mismos archivos

    db.session.expire_all()
    libro = ConciliacionSii.query.one().compras
    reaparecido = next(d for d in libro.documentos if (d.tipo_doc, d.folio) == llave)
    assert reaparecido.aceptado is True
    assert reaparecido.motivo_aceptacion == "Impuesto específico"
    assert reaparecido.aceptado_por_id == usuario_admin.id
    assert libro.n_aceptados == 1


def test_si_el_asiento_se_corrige_la_aceptacion_no_se_arrastra(client, db, empresa, usuario_admin):
    """Aceptada una diferencia, si después el documento pasa a cuadrar, se limpia."""
    login(client, "admin@test.cl")
    _cargar(client)
    conciliacion = ConciliacionSii.query.one()
    documento = _documento_con_diferencia(conciliacion.compras)
    _aceptar(client, conciliacion, documento)

    # El libro corregido: ahora el folio 1001 calza con el SII
    libro_corregido = LIBRO_COMPRAS.replace(
        "<td>90.000</td><td>0</td><td>17.100</td><td>107.100</td>",
        "<td>100.000</td><td>0</td><td>19.000</td><td>119.000</td>",
    )
    _cargar(client, defontana_compra=_archivo(libro_corregido, "compras.xls", "cp1252"))

    db.session.expire_all()
    libro = ConciliacionSii.query.one().compras
    reaparecido = next(d for d in libro.documentos if d.folio == "1001")
    assert reaparecido.estado == "coincide"
    assert reaparecido.aceptado is False
    assert libro.n_aceptados == 0


def test_no_se_acepta_un_documento_que_ya_cuadra(client, db, empresa, usuario_admin):
    login(client, "admin@test.cl")
    _cargar(client)
    conciliacion = ConciliacionSii.query.one()
    cuadrado = next(d for d in conciliacion.ventas.documentos if d.estado == "coincide")

    respuesta = client.post(
        f"/contabilidad/conciliacion-sii/{conciliacion.id}/venta/{cuadrado.id}/aceptar",
        data={}, follow_redirects=True,
    )

    assert "no hay diferencia que aceptar" in respuesta.get_data(as_text=True)
    db.session.expire_all()
    assert db.session.get(ConciliacionSiiDocumento, cuadrado.id).aceptado is False


def test_el_filtro_de_pendientes_ignora_las_aceptadas(client, db, empresa, usuario_admin):
    login(client, "admin@test.cl")
    _cargar(client)
    conciliacion = ConciliacionSii.query.one()
    documento = _documento_con_diferencia(conciliacion.compras)
    _aceptar(client, conciliacion, documento)

    url = f"/contabilidad/conciliacion-sii/{conciliacion.id}/compra"
    pendientes = client.get(f"{url}?filtro=pendientes").get_data(as_text=True)
    aceptadas = client.get(f"{url}?filtro=aceptados").get_data(as_text=True)

    assert "Totales (1 documento)" in pendientes   # sólo queda el "solo en SII"
    assert "Totales (1 documento)" in aceptadas
    assert "✓ aceptada" in aceptadas


def test_sin_permiso_de_edicion_no_se_puede_aceptar(client, db, empresa, usuario_bodega, usuario_admin):
    login(client, "admin@test.cl")
    _cargar(client)
    conciliacion = ConciliacionSii.query.one()
    documento = _documento_con_diferencia(conciliacion.compras)
    client.get("/logout")

    login(client, "bodega@test.cl")
    respuesta = _aceptar(client, conciliacion, documento)

    assert respuesta.status_code == 403
    db.session.expire_all()
    assert db.session.get(ConciliacionSiiDocumento, documento.id).aceptado is False
