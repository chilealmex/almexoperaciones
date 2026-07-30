"""Submódulo Cruce de datos: consistencia de unidad de medida y costo unitario."""

import io

from werkzeug.datastructures import FileStorage

from app.models.conteo_inventario import ItemConteoInventario
from app.utils.importar_conteo import importar_defontana, importar_qms
from tests.conftest import login


CSV_QMS = """﻿Sucursal;Linea Negocio;Categoria; Valor Total Stock CLP ;Stock;Stock Critico;Descripción;Unidad;Código Único;ubicacion_bodega; Valor Unitario CLP
Casa Matriz;GOMAS;CAT-A;100000;10;0;PRODUCTO UNO;UN;COD-001;RACK A1; 10,000
Casa Matriz;ACEROS;CAT-B;70000;7;0;PRODUCTO DOS;RL;COD-002;RACK B2; 10,000
Casa Matriz;GOMAS;CAT-A;30000;3;0;PRODUCTO TRES;UN;COD-003;RACK A2; 10,000
"""

CSV_DEFONTANA = (
    "CodArticulo;Descripci\xf3n Art\xedculo;CodBodega;Nombre Bodega;Saldo Stock;Unidad;Costo Unitario\r\n"
    '"COD-001";"PRODUCTO UNO";"BODEGACENTRAL";"BODEGA CENTRAL";"8";"UN";"9.500"\r\n'
    '"COD-002";"PRODUCTO DOS";"BODEGACENTRAL";"BODEGA CENTRAL";"7";"UN";"10.000"\r\n'
)


def _fs(contenido, nombre):
    return FileStorage(stream=io.BytesIO(contenido), filename=nombre)


def _importar(empresa):
    importar_qms(_fs(CSV_QMS.encode("utf-8"), "qms.csv"), empresa.id)
    importar_defontana(_fs(CSV_DEFONTANA.encode("cp1252"), "def.csv"), empresa.id)


# --- Clasificación del modelo ---


def test_estado_maestro_clasifica_cada_combinacion(db, empresa):
    _importar(empresa)

    dif_costo = ItemConteoInventario.query.filter_by(codigo="COD-001").first()
    assert dif_costo.estado_maestro == "dif_costo"

    dif_unidad = ItemConteoInventario.query.filter_by(codigo="COD-002").first()  # RL vs UN, mismo costo
    assert dif_unidad.estado_maestro == "dif_unidad"

    sin_costo = ItemConteoInventario.query.filter_by(codigo="COD-003").first()  # solo en QMS
    assert sin_costo.estado_maestro == "ok"  # tiene costo (QMS) y no hay unidad Defontana con qué comparar

    dif_costo.costo_unitario_defontana = None
    assert dif_costo.estado_maestro == "ok"
    dif_costo.costo_unitario_defontana = 1
    dif_costo.unidad_defontana = "RL"
    assert dif_costo.estado_maestro == "ambas"


def test_sin_costo_en_ningun_sistema(db, empresa):
    _importar(empresa)
    item = ItemConteoInventario.query.filter_by(codigo="COD-003").first()
    item.costo_unitario_qms = None
    assert item.estado_maestro == "sin_costo"


def test_desviacion_e_impacto_del_costo(db, empresa):
    _importar(empresa)
    item = ItemConteoInventario.query.filter_by(codigo="COD-001").first()

    # QMS 10.000 vs Defontana 9.500 -> +500, sobre 9.500 es ~5.26%
    assert round(item.desviacion_costo_pct, 2) == round(500 / 9500 * 100, 2)
    assert item.impacto_diferencia_costo == 500 * item.cantidad_qms  # 500 * 10 = 5000


def test_sin_datos_de_ambos_costos_no_hay_desviacion(db, empresa):
    _importar(empresa)
    item = ItemConteoInventario.query.filter_by(codigo="COD-003").first()  # sin costo Defontana
    assert item.desviacion_costo_pct is None
    assert item.impacto_diferencia_costo is None


# --- Vista y filtros ---


def test_la_pagina_lista_y_clasifica_los_articulos(client, empresa, usuario_admin):
    _importar(empresa)
    login(client, "admin@test.cl")

    respuesta = client.get("/inventario/cruce-datos")
    assert respuesta.status_code == 200
    cuerpo = respuesta.get_data(as_text=True)
    assert "Cruce de datos" in cuerpo
    assert "COD-001" in cuerpo
    assert "COD-002" in cuerpo


def test_filtro_diferencia_de_costo(client, empresa, usuario_admin):
    _importar(empresa)
    login(client, "admin@test.cl")

    cuerpo = client.get("/inventario/cruce-datos?filtro=dif_costo").get_data(as_text=True)
    assert "COD-001" in cuerpo
    assert "COD-002" not in cuerpo


def test_filtro_distinta_unidad(client, empresa, usuario_admin):
    _importar(empresa)
    login(client, "admin@test.cl")

    cuerpo = client.get("/inventario/cruce-datos?filtro=dif_unidad").get_data(as_text=True)
    assert "COD-002" in cuerpo
    assert "COD-001" not in cuerpo


def test_filtro_sin_costo_cargado(client, empresa, usuario_admin):
    _importar(empresa)
    login(client, "admin@test.cl")
    item = ItemConteoInventario.query.filter_by(codigo="COD-003").first()
    item.costo_unitario_qms = None
    from app.extensions import db
    db.session.commit()

    cuerpo = client.get("/inventario/cruce-datos?filtro=sin_costo").get_data(as_text=True)
    assert "COD-003" in cuerpo
    assert "COD-001" not in cuerpo


def test_exportacion_excel_incluye_desviacion_e_impacto(client, empresa, usuario_admin):
    import io as _io
    from openpyxl import load_workbook

    _importar(empresa)
    login(client, "admin@test.cl")

    respuesta = client.get("/inventario/cruce-datos.xlsx")
    assert respuesta.status_code == 200
    hoja = load_workbook(_io.BytesIO(respuesta.get_data())).active
    titulos = [hoja.cell(row=4, column=i).value for i in range(1, hoja.max_column + 1)]
    assert "Desviación costo (%)" in titulos
    assert "Impacto en stock QMS" in titulos


def test_no_se_solapa_con_ajuste_inventario(client, empresa, usuario_admin):
    """Ajuste inventario ya no debe mostrar columnas de unidad ni filtros de calidad del maestro."""
    _importar(empresa)
    login(client, "admin@test.cl")

    cuerpo = client.get("/inventario/ajuste").get_data(as_text=True)
    assert "Un. QMS" not in cuerpo
    assert "Diferencia de costo" not in cuerpo
    assert "Distinta unidad" not in cuerpo
    assert "Ir a Cruce de datos" in cuerpo  # enlace cruzado entre ambos submódulos


def test_cruce_de_datos_exige_permiso(client, empresa, usuario_bodega):
    login(client, "bodega@test.cl")
    assert client.get("/inventario/cruce-datos").status_code == 200


def test_cruce_de_datos_exige_sesion(client, empresa):
    respuesta = client.get("/inventario/cruce-datos")
    assert respuesta.status_code == 302
    assert "/login" in respuesta.headers["Location"]
