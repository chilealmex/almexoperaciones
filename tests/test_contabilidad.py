"""Submódulo Contabilidad > Provisión de Ingresos: importar acumulando y editar 4 columnas."""

import io
from datetime import date

from openpyxl import Workbook

from app.extensions import db as _db
from app.models.contabilidad import ProvisionIngreso
from app.utils.provision_ingresos_excel import PlanillaInvalida, leer_provisiones, mes_legible
from tests.conftest import login

TITULOS = [
    "Mes.año", "Cbte Prov", "OT", "Monto Provisión", "Reversa", "Mes Reversa",
    "Cbte Reversa", "Cliente", "Centro de Costos", "Rut", "Obs", "Saldo",
]


def _planilla(filas, hoja="Control"):
    """Arma un xlsx en memoria con la forma de la planilla real: totales, títulos y datos."""
    libro = Workbook()
    ws = libro.active
    ws.title = hoja
    ws.append([None, None, None, 0, 0, None, None, None, None, None, None, 0])  # fila de totales
    ws.append(TITULOS)
    for fila in filas:
        ws.append(fila)
    memoria = io.BytesIO()
    libro.save(memoria)
    memoria.seek(0)
    return memoria


FILA_1 = [date(2026, 3, 1), 67, 6095, 3700000, 3700000, date(2026, 5, 1), 109,
          "CIA MINERA COLLAHUASI", "EMPNEGVTAVTAPRE", "89468900-5", None, 0]
FILA_2 = [date(2026, 3, 1), 67, 6119, 91035000, None, None, None,
          "MINERA LOS PELAMBRES", "EMPNEGVTAVTAOEM", "96790240-3", None, 91035000]
FILA_3 = [date(2026, 4, 1), 114, 6136, 290000, None, None, None,
          "CONTITECH CHILE S.A.", "EMPNEGVTAVTAOEM", "83070800-6", None, 290000]


# --- Lectura de la planilla ---------------------------------------------


def test_lee_las_lineas_de_la_hoja_control():
    lineas = leer_provisiones(_planilla([FILA_1, FILA_2]))
    assert len(lineas) == 2
    assert lineas[0]["cbte_prov"] == "67"
    assert lineas[0]["ot"] == "6095"
    assert lineas[0]["monto_provision"] == 3_700_000
    assert lineas[0]["mes_reversa"] == "05.2026"
    assert lineas[1]["reversa"] is None
    assert lineas[1]["saldo"] == 91_035_000


def test_el_mes_de_reversa_admite_texto_libre():
    """En la planilla real hay una línea reversada en dos meses: 'may-26 y jun-26'."""
    fila = list(FILA_1)
    fila[5] = "may-26 y jun-26"
    fila[6] = "456-204"
    lineas = leer_provisiones(_planilla([fila]))
    assert lineas[0]["mes_reversa"] == "may-26 y jun-26"
    assert lineas[0]["cbte_reversa"] == "456-204"


def test_se_ignoran_las_filas_sin_los_datos_que_identifican_la_linea():
    lineas = leer_provisiones(_planilla([FILA_1, [None, None, None, 999, None, None, None, None, None, None, None, 0]]))
    assert len(lineas) == 1


def test_un_archivo_sin_la_hoja_control_avisa_en_vez_de_reventar():
    try:
        leer_provisiones(_planilla([FILA_1], hoja="Otra"))
        assert False, "debería haber avisado que falta la hoja Control"
    except PlanillaInvalida as error:
        assert "Control" in str(error)


def test_mes_legible():
    assert mes_legible(date(2026, 5, 1)) == "may-26"
    assert mes_legible(None) == ""


# --- Permisos ------------------------------------------------------------


def test_bodega_no_puede_ver_contabilidad(client, usuario_bodega):
    login(client, "bodega@test.cl")
    assert client.get("/contabilidad/provision-ingresos").status_code == 403


# --- Importar acumulando -------------------------------------------------


def test_importar_carga_las_lineas_de_la_planilla(client, usuario_admin, empresa, db):
    login(client, "admin@test.cl")
    respuesta = client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_1, FILA_2]), "provision.xlsx")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert respuesta.status_code == 200
    assert ProvisionIngreso.query.filter_by(empresa_id=empresa.id).count() == 2


def test_reimportar_no_duplica_ni_borra_lo_anterior(client, usuario_admin, empresa, db):
    """Lo pedido: cada importación agrega lo nuevo y deja lo que ya estaba."""
    login(client, "admin@test.cl")
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_1, FILA_2]), "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    # La segunda planilla repite las dos anteriores y trae una tercera.
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_1, FILA_2, FILA_3]), "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    lineas = ProvisionIngreso.query.filter_by(empresa_id=empresa.id).all()
    assert len(lineas) == 3
    assert {l.ot for l in lineas} == {"6095", "6119", "6136"}


def test_reimportar_respeta_lo_editado_a_mano(client, usuario_admin, empresa, db):
    login(client, "admin@test.cl")
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_2]), "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    linea = ProvisionIngreso.query.filter_by(empresa_id=empresa.id).one()
    linea.reversa = 91_035_000
    linea.mes_reversa = "07.2026"
    linea.saldo = 0
    _db.session.commit()

    client.post(  # se vuelve a subir el mismo Excel, que trae la línea sin reversar
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_2]), "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    _db.session.refresh(linea)
    assert linea.reversa == 91_035_000  # no se pisó con lo del archivo
    assert linea.mes_reversa == "07.2026"
    assert linea.saldo == 0


def test_importar_un_archivo_que_no_es_la_planilla_avisa(client, usuario_admin, empresa, db):
    login(client, "admin@test.cl")
    respuesta = client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_1], hoja="Hoja1"), "otra.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    assert "Control" in respuesta.get_data(as_text=True)
    assert ProvisionIngreso.query.count() == 0


# --- Edición de las 4 columnas ------------------------------------------


def test_solo_se_editan_reversa_mes_reversa_cbte_reversa_y_saldo(client, usuario_admin, empresa, db):
    login(client, "admin@test.cl")
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_2]), "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    linea = ProvisionIngreso.query.filter_by(empresa_id=empresa.id).one()

    client.post(
        "/contabilidad/provision-ingresos/guardar",
        data={
            f"linea-{linea.id}-reversa": "$91.035.000",
            f"linea-{linea.id}-mes_reversa": "may-26 y jun-26",
            f"linea-{linea.id}-cbte_reversa": "456-204",
            # Estos no son editables: aunque se envíen, deben ignorarse.
            f"linea-{linea.id}-monto_provision": "1",
            f"linea-{linea.id}-cliente": "OTRO CLIENTE",
        },
        follow_redirects=True,
    )
    _db.session.refresh(linea)
    assert linea.reversa == 91_035_000
    assert linea.mes_reversa == "may-26 y jun-26"
    assert linea.cbte_reversa == "456-204"
    assert linea.saldo == 0
    assert linea.monto_provision == 91_035_000  # intacto
    assert linea.cliente == "MINERA LOS PELAMBRES"  # intacto


def test_dejar_la_reversa_vacia_la_borra(client, usuario_admin, empresa, db):
    """En una línea abierta se puede borrar la reversa; el saldo vuelve a la provisión."""
    login(client, "admin@test.cl")
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_2]), "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    linea = ProvisionIngreso.query.filter_by(empresa_id=empresa.id).one()
    client.post(
        "/contabilidad/provision-ingresos/guardar",
        data={f"linea-{linea.id}-reversa": "40000000", f"linea-{linea.id}-mes_reversa": "06.2026",
              f"linea-{linea.id}-cbte_reversa": "456"},
        follow_redirects=True,
    )
    _db.session.refresh(linea)
    assert linea.reversa == 40_000_000

    client.post(
        "/contabilidad/provision-ingresos/guardar",
        data={f"linea-{linea.id}-reversa": "", f"linea-{linea.id}-mes_reversa": "",
              f"linea-{linea.id}-cbte_reversa": ""},
        follow_redirects=True,
    )
    _db.session.refresh(linea)
    assert linea.reversa is None
    assert linea.mes_reversa is None
    assert linea.saldo == 91_035_000   # vuelve a quedar toda la provisión pendiente


# --- Pantalla ------------------------------------------------------------


def test_la_pantalla_muestra_las_lineas_y_los_totales(client, usuario_admin, empresa, db):
    login(client, "admin@test.cl")
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_1, FILA_2]), "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    texto = client.get("/contabilidad/provision-ingresos").get_data(as_text=True)
    assert "MINERA LOS PELAMBRES" in texto
    assert "$94.735.000" in texto  # monto provisión total
    assert "$91.035.000" in texto  # saldo total
    assert "Marzo" in texto


def test_el_mes_y_el_ano_se_muestran_en_columnas_separadas(client, usuario_admin, empresa, db):
    login(client, "admin@test.cl")
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_1]), "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    texto = client.get("/contabilidad/provision-ingresos").get_data(as_text=True)
    assert "<th>Mes</th><th>Año</th>" in texto
    assert "Marzo" in texto
    assert "2026" in texto


def test_se_puede_filtrar_por_mes_y_por_ano_por_separado(client, usuario_admin, empresa, db):
    login(client, "admin@test.cl")
    otro_ano = list(FILA_1)
    otro_ano[0] = date(2025, 4, 1)
    otro_ano[2] = 9999  # otra OT, para que sea una línea distinta
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_1, FILA_2, FILA_3, otro_ano]), "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    solo_abril = client.get("/contabilidad/provision-ingresos?mes=4").get_data(as_text=True)
    assert "CONTITECH" in solo_abril      # abril 2026
    assert "COLLAHUASI" in solo_abril     # abril 2025, mismo mes distinto año
    assert "PELAMBRES" not in solo_abril  # marzo

    solo_2025 = client.get("/contabilidad/provision-ingresos?anio=2025").get_data(as_text=True)
    assert "9999" in solo_2025
    assert "CONTITECH" not in solo_2025

    abril_2026 = client.get("/contabilidad/provision-ingresos?mes=4&anio=2026").get_data(as_text=True)
    assert "CONTITECH" in abril_2026
    assert "9999" not in abril_2026


def test_la_linea_con_saldo_cero_aparece_como_cerrada(client, usuario_admin, empresa, db):
    login(client, "admin@test.cl")
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_1, FILA_2]), "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    texto = client.get("/contabilidad/provision-ingresos").get_data(as_text=True)
    assert "Cerrado" in texto    # FILA_1 tiene saldo 0
    assert "Pendiente" in texto  # FILA_2 tiene saldo 91.035.000


def test_se_puede_filtrar_por_saldo_mayor_a_cero_y_por_cerradas(client, usuario_admin, empresa, db):
    login(client, "admin@test.cl")
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_1, FILA_2, FILA_3]), "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    pendientes = client.get("/contabilidad/provision-ingresos?estado=pendiente").get_data(as_text=True)
    assert "PELAMBRES" in pendientes    # saldo 91.035.000
    assert "CONTITECH" in pendientes    # saldo 290.000
    assert "COLLAHUASI" not in pendientes  # saldo 0, está cerrada

    cerradas = client.get("/contabilidad/provision-ingresos?estado=cerrado").get_data(as_text=True)
    assert "COLLAHUASI" in cerradas
    assert "PELAMBRES" not in cerradas


def test_se_puede_eliminar_una_linea(client, usuario_admin, empresa, db):
    login(client, "admin@test.cl")
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_1]), "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    linea = ProvisionIngreso.query.filter_by(empresa_id=empresa.id).one()
    client.post(f"/contabilidad/provision-ingresos/{linea.id}/eliminar", follow_redirects=True)
    assert ProvisionIngreso.query.count() == 0


def test_exportar_devuelve_un_excel(client, usuario_admin, empresa, db):
    login(client, "admin@test.cl")
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_1]), "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    respuesta = client.get("/contabilidad/provision-ingresos/exportar.xlsx")
    assert respuesta.status_code == 200
    assert "spreadsheetml" in respuesta.headers["Content-Type"]


def test_eliminar_una_linea_avisa_antes_y_dice_que_se_va_a_perder(client, usuario_admin, empresa, db):
    """El botón de eliminar tiene que pedir confirmación explicando qué línea es."""
    login(client, "admin@test.cl")
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_1]), "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    texto = client.get("/contabilidad/provision-ingresos").get_data(as_text=True)

    assert "Se eliminar\\u00e1 esta l\\u00ednea de la provisi\\u00f3n" in texto
    assert "no se puede deshacer" in texto
    assert "COLLAHUASI" in texto           # dice de qué línea se trata
    assert "Tiene reversa registrada" in texto  # y que se pierde la reversa


def test_el_aviso_de_borrado_no_rompe_el_atributo_aunque_el_cliente_traiga_comillas(client, usuario_admin, empresa, db):
    """El mensaje viaja dentro de un atributo HTML: si se corta, el borrado ocurriría sin preguntar."""
    login(client, "admin@test.cl")
    con_comillas = list(FILA_1)
    con_comillas[7] = 'CLIENTE "RARO" S.A.'
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([con_comillas]), "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    texto = client.get("/contabilidad/provision-ingresos").get_data(as_text=True)

    # El atributo va con comillas simples y el contenido no puede traer una comilla
    # simple sin escapar, ni una comilla doble cruda que lo corte antes de tiempo.
    inicio = texto.index("onclick='return confirm(")
    fin = texto.index("' title=\"Eliminar\"", inicio)
    atributo = texto[inicio + len("onclick='"):fin]
    assert atributo.startswith("return confirm(")
    assert atributo.endswith(")")
    assert "'" not in atributo, "una comilla simple sin escapar cortaría el atributo"


# --- Saldo calculado y línea cerrada -------------------------------------


def test_el_saldo_sale_de_restar_la_reversa_a_la_provision(client, usuario_admin, empresa, db):
    login(client, "admin@test.cl")
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_2]), "provision.xlsx")},   # provisión 91.035.000, sin reversa
        content_type="multipart/form-data", follow_redirects=True,
    )
    linea = ProvisionIngreso.query.filter_by(empresa_id=empresa.id).one()
    assert linea.saldo == 91_035_000

    client.post(
        "/contabilidad/provision-ingresos/guardar",
        data={f"linea-{linea.id}-reversa": "$40.000.000"},
        follow_redirects=True,
    )
    _db.session.refresh(linea)
    assert linea.saldo == 91_035_000 - 40_000_000  # no se escribe: se calcula


def test_al_reversar_todo_el_saldo_queda_en_cero_y_la_linea_se_cierra(client, usuario_admin, empresa, db):
    login(client, "admin@test.cl")
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_2]), "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    linea = ProvisionIngreso.query.filter_by(empresa_id=empresa.id).one()
    client.post(
        "/contabilidad/provision-ingresos/guardar",
        data={f"linea-{linea.id}-reversa": "91035000"},
        follow_redirects=True,
    )
    _db.session.refresh(linea)
    assert linea.saldo == 0

    texto = client.get("/contabilidad/provision-ingresos").get_data(as_text=True)
    assert "saldo-cerrado" in texto          # la celda va en verde
    assert "Cerrado" in texto


def test_una_linea_cerrada_no_se_puede_modificar(client, usuario_admin, empresa, db):
    """Saldo en $0: queda con candado y el servidor ignora cualquier cambio."""
    login(client, "admin@test.cl")
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_1]), "provision.xlsx")},   # ya viene reversada, saldo 0
        content_type="multipart/form-data", follow_redirects=True,
    )
    linea = ProvisionIngreso.query.filter_by(empresa_id=empresa.id).one()
    assert linea.saldo == 0

    client.post(
        "/contabilidad/provision-ingresos/guardar",
        data={f"linea-{linea.id}-reversa": "1", f"linea-{linea.id}-cbte_reversa": "PISADO"},
        follow_redirects=True,
    )
    _db.session.refresh(linea)
    assert linea.reversa == 3_700_000    # intacta
    assert linea.cbte_reversa == "109"
    assert linea.saldo == 0


def test_un_superadmin_si_puede_corregir_una_linea_cerrada(client, empresa, db):
    """Hace falta para arreglar un monto mal escrito que dejó la línea en $0."""
    from tests.test_permissions import _crear_superadmin

    _crear_superadmin(db, empresa)
    login(client, "super@test.cl")
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_1]), "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    linea = ProvisionIngreso.query.filter_by(empresa_id=empresa.id).one()
    assert linea.saldo == 0

    client.post(
        "/contabilidad/provision-ingresos/guardar",
        data={f"linea-{linea.id}-reversa": "1000000"},
        follow_redirects=True,
    )
    _db.session.refresh(linea)
    assert linea.reversa == 1_000_000
    assert linea.saldo == 3_700_000 - 1_000_000


def test_la_columna_se_llama_monto_reversa_y_el_saldo_ya_no_se_escribe(client, usuario_admin, empresa, db):
    login(client, "admin@test.cl")
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_2]), "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    linea = ProvisionIngreso.query.filter_by(empresa_id=empresa.id).one()
    texto = client.get("/contabilidad/provision-ingresos").get_data(as_text=True)

    assert "<th class=\"num\">Monto Reversa</th>" in texto
    assert f'name="linea-{linea.id}-saldo"' not in texto   # el saldo dejó de ser un campo
    assert f'id="saldo-{linea.id}"' in texto               # ahora es una celda calculada
    assert "data-confirmar-cambio" in texto                # y avisa antes de pisar un dato


# --- Plantilla descargable ----------------------------------------------


def test_la_plantilla_se_descarga_con_las_columnas_de_la_planilla(client, usuario_admin, empresa, db):
    login(client, "admin@test.cl")
    respuesta = client.get("/contabilidad/provision-ingresos/plantilla.xlsx")
    assert respuesta.status_code == 200
    assert "spreadsheetml" in respuesta.headers["Content-Type"]

    from openpyxl import load_workbook
    hoja = load_workbook(io.BytesIO(respuesta.data))["Control"]
    titulos = [c.value for c in hoja[1] if c.value]
    assert titulos == [
        "Mes.año", "Cbte Prov", "OT", "Monto Provisión", "Reversa", "Mes Reversa",
        "Cbte Reversa", "Cliente", "Centro de Costos", "Rut", "Obs", "Saldo",
    ]


def test_la_plantilla_descargada_se_puede_llenar_y_volver_a_subir(client, usuario_admin, empresa, db):
    """El circuito completo: bajar la plantilla, escribir en ella y cargarla."""
    login(client, "admin@test.cl")
    from openpyxl import load_workbook

    descargada = client.get("/contabilidad/provision-ingresos/plantilla.xlsx").data
    libro = load_workbook(io.BytesIO(descargada))
    hoja = libro["Control"]
    hoja.delete_rows(2)  # se borra la fila de ejemplo, como haría cualquiera
    hoja.append([date(2026, 7, 1), 900, 7001, 5_000_000, None, None, None,
                 "CLIENTE DE PRUEBA", "CENTRO-1", "76000000-1", None, 5_000_000])
    memoria = io.BytesIO()
    libro.save(memoria)
    memoria.seek(0)

    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (memoria, "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    linea = ProvisionIngreso.query.filter_by(empresa_id=empresa.id).one()
    assert linea.cliente == "CLIENTE DE PRUEBA"
    assert linea.monto_provision == 5_000_000
    assert linea.saldo == 5_000_000
    assert linea.ot == "7001"


def test_se_siguen_leyendo_las_planillas_con_los_titulos_en_la_segunda_fila(client, usuario_admin, empresa, db):
    """La planilla original trae una fila de totales encima de los títulos."""
    login(client, "admin@test.cl")
    client.post(
        "/contabilidad/provision-ingresos/importar",
        data={"archivo": (_planilla([FILA_1, FILA_2]), "provision.xlsx")},
        content_type="multipart/form-data", follow_redirects=True,
    )
    assert ProvisionIngreso.query.filter_by(empresa_id=empresa.id).count() == 2


def test_la_pantalla_ofrece_descargar_la_plantilla(client, usuario_admin, empresa, db):
    login(client, "admin@test.cl")
    texto = client.get("/contabilidad/provision-ingresos").get_data(as_text=True)
    assert "Descargar plantilla" in texto
