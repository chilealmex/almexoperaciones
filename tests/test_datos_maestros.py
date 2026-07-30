"""Módulo Datos maestros: clientes, proveedores e importación masiva."""

import io

from openpyxl import Workbook, load_workbook

from app.models.cliente import Cliente, Proveedor
from app.models.permiso import MODULOS
from tests.conftest import login


def _archivo(wb, nombre):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return (buf, nombre)


def test_datos_maestros_esta_en_la_matriz_de_permisos():
    assert "datos_maestros" in MODULOS


def test_resumen_muestra_conteos(client, empresa, usuario_admin, db):
    db.session.add(Cliente(empresa_id=empresa.id, rut="11.111.111-1", razon_social="Cliente Uno"))
    db.session.commit()
    login(client, "admin@test.cl")

    respuesta = client.get("/datos-maestros/")
    assert respuesta.status_code == 200
    assert "Datos maestros" in respuesta.get_data(as_text=True)


def test_crear_editar_y_listar_cliente(client, empresa, usuario_admin):
    login(client, "admin@test.cl")

    respuesta = client.post(
        "/datos-maestros/clientes/nuevo",
        data={"rut": "76.123.456-0", "razon_social": "Cliente Nuevo SPA", "activo": "y"},
        follow_redirects=True,
    )
    assert respuesta.status_code == 200
    cliente = Cliente.query.filter_by(rut="76.123.456-0").first()
    assert cliente is not None

    respuesta = client.get("/datos-maestros/clientes")
    assert "Cliente Nuevo SPA" in respuesta.get_data(as_text=True)

    respuesta = client.post(
        f"/datos-maestros/clientes/{cliente.id}/editar",
        data={"rut": "76.123.456-0", "razon_social": "Cliente Editado SPA", "activo": "y"},
        follow_redirects=True,
    )
    assert "Cliente Editado SPA" in respuesta.get_data(as_text=True)


def test_crear_proveedor(client, empresa, usuario_admin):
    login(client, "admin@test.cl")

    respuesta = client.post(
        "/datos-maestros/proveedores/nuevo",
        data={"rut": "96.789.123-1", "razon_social": "Proveedor Nuevo Ltda", "activo": "y"},
        follow_redirects=True,
    )
    assert respuesta.status_code == 200
    assert Proveedor.query.filter_by(rut="96.789.123-1").first() is not None


def test_exportacion_excel_clientes_y_proveedores(client, empresa, usuario_admin, db):
    db.session.add(Cliente(empresa_id=empresa.id, rut="11.111.111-1", razon_social="Cliente Uno"))
    db.session.add(Proveedor(empresa_id=empresa.id, rut="22.222.222-2", razon_social="Proveedor Uno"))
    db.session.commit()
    login(client, "admin@test.cl")

    for url in ("/datos-maestros/clientes.xlsx", "/datos-maestros/proveedores.xlsx"):
        respuesta = client.get(url)
        assert respuesta.status_code == 200
        assert "spreadsheetml" in respuesta.headers["Content-Type"]


def test_plantillas_descargables(client, empresa, usuario_admin):
    login(client, "admin@test.cl")
    for url in ("/datos-maestros/clientes/plantilla.xlsx", "/datos-maestros/proveedores/plantilla.xlsx"):
        respuesta = client.get(url)
        assert respuesta.status_code == 200
        hoja = load_workbook(io.BytesIO(respuesta.get_data())).active
        assert hoja.cell(row=4, column=1).value == "RUT"


def test_importar_clientes_crea_y_actualiza_por_rut(client, empresa, usuario_admin):
    login(client, "admin@test.cl")

    wb = Workbook()
    ws = wb.active
    ws.append(["RUT", "Razón social", "Email", "Activo"])
    ws.append(["76.123.456-0", "Cliente Importado SPA", "contacto@cliente.cl", "SI"])
    ws.append(["11111111", "Fila sin razón social ni RUT válido", "", "SI"])  # RUT inválido: sin DV

    respuesta = client.post(
        "/datos-maestros/clientes/importar",
        data={"cli-archivo": _archivo(wb, "clientes.xlsx")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert respuesta.status_code == 200
    assert Cliente.query.filter_by(rut="76.123.456-0").first() is not None
    assert Cliente.query.count() == 1  # la fila inválida se descarta

    # reimportar el mismo RUT actualiza en vez de duplicar
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(["RUT", "Razón social"])
    ws2.append(["76.123.456-0", "Cliente Actualizado SPA"])
    client.post(
        "/datos-maestros/clientes/importar",
        data={"cli-archivo": _archivo(wb2, "clientes2.xlsx")},
        content_type="multipart/form-data",
    )
    assert Cliente.query.count() == 1
    assert Cliente.query.first().razon_social == "Cliente Actualizado SPA"


def test_importar_proveedores(client, empresa, usuario_admin):
    login(client, "admin@test.cl")

    wb = Workbook()
    ws = wb.active
    ws.append(["RUT", "Razón social"])
    ws.append(["96.789.123-1", "Proveedor Importado Ltda"])

    respuesta = client.post(
        "/datos-maestros/proveedores/importar",
        data={"prov-archivo": _archivo(wb, "proveedores.xlsx")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert respuesta.status_code == 200
    assert Proveedor.query.filter_by(rut="96.789.123-1").first() is not None


def test_datos_maestros_exige_permiso(client, empresa, usuario_bodega):
    """Bodega solo tiene permiso sobre inventario, no sobre datos_maestros."""
    login(client, "bodega@test.cl")
    assert client.get("/datos-maestros/").status_code == 403
    assert client.get("/datos-maestros/clientes").status_code == 403


def test_datos_maestros_exige_sesion(client, empresa):
    respuesta = client.get("/datos-maestros/clientes")
    assert respuesta.status_code == 302
    assert "/login" in respuesta.headers["Location"]


def test_rutas_viejas_de_clientes_y_proveedores_ya_no_existen(client, empresa, usuario_admin):
    """Se movieron a datos_maestros; las rutas antiguas no deben quedar registradas."""
    login(client, "admin@test.cl")
    assert client.get("/contratos/clientes").status_code == 404
    assert client.get("/arriendos/proveedores").status_code == 404
