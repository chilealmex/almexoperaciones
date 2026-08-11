import json

from app.models.conteo_inventario import ItemConteoInventario
from tests.conftest import login


def _item(db, empresa, codigo, qms, defontana, fisica=None):
    item = ItemConteoInventario(
        empresa_id=empresa.id,
        codigo=codigo,
        nombre=f"Artículo {codigo}",
        cantidad_qms=qms,
        cantidad_defontana=defontana,
        cantidad_fisica=fisica,
    )
    db.session.add(item)
    db.session.commit()
    return item


def test_stock_muestra_cantidades_y_diferencias(client, db, empresa, usuario_admin):
    _item(db, empresa, "COD-A", 10, 4)
    login(client, "admin@test.cl")

    body = client.get("/inventario/stock").get_data(as_text=True)
    assert "COD-A" in body
    assert "+6" in body  # diferencia entre sistemas visible en la misma fila


def test_filtro_diferencias_coincide_con_el_modelo(client, db, empresa, usuario_admin):
    _item(db, empresa, "IGUALES", 5, 5)  # sin diferencia
    _item(db, empresa, "DESCUADRE", 9, 2)  # sistemas no cuadran
    _item(db, empresa, "FISICO-DISTINTO", 3, 3, fisica=8)  # físico no coincide
    _item(db, empresa, "TODO-OK", 4, 4, fisica=4)  # los tres cuadran
    login(client, "admin@test.cl")

    body = client.get("/inventario/stock?filtro=diferencias").get_data(as_text=True)
    assert "DESCUADRE" in body
    assert "FISICO-DISTINTO" in body
    assert "IGUALES" not in body
    assert "TODO-OK" not in body

    esperados = {i.codigo for i in ItemConteoInventario.query.all() if i.tiene_diferencia}
    assert esperados == {"DESCUADRE", "FISICO-DISTINTO"}


def test_filtros_sin_contar_y_contados(client, db, empresa, usuario_admin):
    _item(db, empresa, "PENDIENTE", 5, 5)
    _item(db, empresa, "YA-CONTADO", 5, 5, fisica=5)
    login(client, "admin@test.cl")

    sin_contar = client.get("/inventario/stock?filtro=sin_contar").get_data(as_text=True)
    assert "PENDIENTE" in sin_contar and "YA-CONTADO" not in sin_contar

    contados = client.get("/inventario/stock?filtro=contados").get_data(as_text=True)
    assert "YA-CONTADO" in contados and "PENDIENTE" not in contados


def test_buscar_por_codigo(client, db, empresa, usuario_admin):
    _item(db, empresa, "TORNILLO-1", 1, 1)
    _item(db, empresa, "TUERCA-9", 2, 2)
    login(client, "admin@test.cl")

    body = client.get("/inventario/stock?q=tornillo").get_data(as_text=True)
    assert "TORNILLO-1" in body and "TUERCA-9" not in body


def test_contar_en_la_misma_fila_guarda_y_devuelve_diferencias(client, db, empresa, usuario_admin):
    item = _item(db, empresa, "COD-B", 10, 8)
    login(client, "admin@test.cl")

    r = client.post(
        f"/inventario/stock/{item.id}/contar",
        data=json.dumps({"cantidad": "9"}),
        content_type="application/json",
    )
    assert r.status_code == 200
    datos = r.get_json()
    assert datos["ok"] is True
    assert datos["dif_qms"] == -1
    assert datos["dif_defontana"] == 1

    db.session.refresh(item)
    assert item.cantidad_fisica == 9
    assert item.contado_por_id == usuario_admin.id
    assert item.contado_en is not None


def test_vaciar_el_campo_borra_el_conteo(client, db, empresa, usuario_admin):
    item = _item(db, empresa, "COD-C", 5, 5, fisica=3)
    login(client, "admin@test.cl")

    r = client.post(
        f"/inventario/stock/{item.id}/contar",
        data=json.dumps({"cantidad": ""}),
        content_type="application/json",
    )
    assert r.status_code == 200

    db.session.refresh(item)
    assert item.cantidad_fisica is None
    assert not item.contado


def test_cantidad_invalida_es_rechazada(client, db, empresa, usuario_admin):
    item = _item(db, empresa, "COD-D", 5, 5)
    login(client, "admin@test.cl")

    for valor in ["abc", "-3"]:
        r = client.post(
            f"/inventario/stock/{item.id}/contar",
            data=json.dumps({"cantidad": valor}),
            content_type="application/json",
        )
        assert r.status_code == 400
        assert r.get_json()["ok"] is False

    db.session.refresh(item)
    assert item.cantidad_fisica is None


def test_usuario_sin_permiso_editar_no_puede_contar(client, db, empresa, usuario_bodega):
    item = _item(db, empresa, "COD-E", 5, 5)
    login(client, "bodega@test.cl")

    r = client.post(
        f"/inventario/stock/{item.id}/contar",
        data=json.dumps({"cantidad": "4"}),
        content_type="application/json",
    )
    assert r.status_code == 403

    db.session.refresh(item)
    assert item.cantidad_fisica is None


# --- Al entrar se muestra lo que falta por contar ---

def _dos_articulos(empresa, db):
    db.session.add_all([
        ItemConteoInventario(empresa_id=empresa.id, codigo="POR-CONTAR", nombre="Pendiente",
                             cantidad_qms=5, cantidad_defontana=5),
        ItemConteoInventario(empresa_id=empresa.id, codigo="YA-CONTADO", nombre="Listo",
                             cantidad_qms=5, cantidad_defontana=5, cantidad_fisica=5),
    ])
    db.session.commit()


def test_al_entrar_se_ven_solo_los_que_faltan_por_contar(client, usuario_admin, empresa, db):
    """Durante la toma, lo ya contado sólo estorba: había que filtrar en cada recarga."""
    _dos_articulos(empresa, db)
    login(client, "admin@test.cl")

    texto = client.get("/inventario/stock").get_data(as_text=True)
    assert "POR-CONTAR" in texto
    assert "YA-CONTADO" not in texto


def test_se_pueden_ver_todos_a_proposito(client, usuario_admin, empresa, db):
    _dos_articulos(empresa, db)
    login(client, "admin@test.cl")

    texto = client.get("/inventario/stock?filtro=todos").get_data(as_text=True)
    assert "POR-CONTAR" in texto
    assert "YA-CONTADO" in texto


def test_el_filtro_de_contados_sigue_funcionando(client, usuario_admin, empresa, db):
    _dos_articulos(empresa, db)
    login(client, "admin@test.cl")

    texto = client.get("/inventario/stock?filtro=contados").get_data(as_text=True)
    assert "YA-CONTADO" in texto
    assert "POR-CONTAR" not in texto


def test_un_filtro_inventado_cae_en_lo_pendiente_y_no_revienta(client, usuario_admin, empresa, db):
    _dos_articulos(empresa, db)
    login(client, "admin@test.cl")

    respuesta = client.get("/inventario/stock?filtro=cualquier-cosa")
    assert respuesta.status_code == 200
    assert "POR-CONTAR" in respuesta.get_data(as_text=True)


def test_la_tabla_deja_las_filas_todas_del_mismo_alto(client, usuario_admin, empresa, db):
    """Nombre, ubicación y línea en una sola línea; el texto completo va en title."""
    db.session.add(ItemConteoInventario(
        empresa_id=empresa.id, codigo="LARGO-01",
        nombre="TUERCAS DE ANCLAJE DE ACERO INOXIDABLE 3/8 CON ARANDELA",
        cantidad_qms=1, cantidad_defontana=1,
        ubicacion="RACK PRINCIPAL PASILLO 3 NIVEL 2", linea_negocio="REPUESTOS INDUSTRIALES"))
    db.session.commit()

    login(client, "admin@test.cl")
    texto = client.get("/inventario/stock").get_data(as_text=True)

    assert "tabla-stock" in texto
    assert 'title="TUERCAS DE ANCLAJE DE ACERO INOXIDABLE 3/8 CON ARANDELA"' in texto
    assert "celda-ubicacion" in texto and "celda-linea" in texto


def test_el_excel_descarga_todo_aunque_la_pantalla_muestre_solo_lo_pendiente(client, usuario_admin, empresa, db):
    """Un archivo descargado se entiende como el registro completo.

    La pantalla parte mostrando lo que falta por contar, pero entregar un Excel
    recortado sin avisar se presta a confusión: quien lo abre cree que ese es
    todo el inventario.
    """
    from openpyxl import load_workbook
    import io

    _dos_articulos(empresa, db)
    login(client, "admin@test.cl")

    hoja = load_workbook(io.BytesIO(client.get("/inventario/stock.xlsx").data)).active
    codigos = {fila[0] for fila in hoja.iter_rows(min_row=2, values_only=True) if fila and fila[0]}
    assert {"POR-CONTAR", "YA-CONTADO"} <= codigos


def test_si_se_elige_un_filtro_el_excel_lo_respeta(client, usuario_admin, empresa, db):
    from openpyxl import load_workbook
    import io

    _dos_articulos(empresa, db)
    login(client, "admin@test.cl")

    hoja = load_workbook(io.BytesIO(client.get("/inventario/stock.xlsx?filtro=sin_contar").data)).active
    codigos = {fila[0] for fila in hoja.iter_rows(min_row=2, values_only=True) if fila and fila[0]}
    assert "POR-CONTAR" in codigos
    assert "YA-CONTADO" not in codigos
