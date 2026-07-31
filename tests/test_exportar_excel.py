"""Los informes en Excel deben salir con datos, formato y totales correctos."""

import io
from datetime import date, timedelta

from openpyxl import load_workbook

from app.extensions import db as _db
from app.models.activo_fijo import ActivoFijo
from app.models.cliente import Cliente
from app.models.contrato import ContratoCliente
from app.models.conteo_inventario import ItemConteoInventario
from app.utils.exportar import CLP, ENTERO, col, construir_libro
from tests.conftest import login


def _hoja(respuesta):
    return load_workbook(io.BytesIO(respuesta.get_data())).active


# --- Generador de libros ---


def test_el_libro_lleva_encabezado_titulos_y_totales():
    columnas = [
        col("Código", ancho=12, total="texto"),
        col("Cantidad", ancho=10, formato=ENTERO, total="suma"),
        col("Valor", ancho=14, formato=CLP, total="suma"),
    ]
    filas = [["A-1", 3, 1500], ["A-2", 7, 2500]]

    hoja = load_workbook(construir_libro("Prueba", columnas, filas, "Filtro X")).active

    assert hoja.title == "Prueba"
    assert hoja["A1"].value == "Almex Operaciones — Prueba"
    assert "Filtro X" in hoja["A2"].value
    assert [hoja.cell(row=4, column=i).value for i in (1, 2, 3)] == ["Código", "Cantidad", "Valor"]
    assert hoja.cell(row=5, column=1).value == "A-1"

    fila_totales = hoja.max_row
    assert hoja.cell(row=fila_totales, column=1).value == "Totales (2)"
    assert hoja.cell(row=fila_totales, column=2).value == 10
    assert hoja.cell(row=fila_totales, column=3).value == 4000
    assert hoja.cell(row=5, column=3).number_format == CLP
    assert hoja.freeze_panes == "A5"  # los títulos quedan fijos al hacer scroll


def test_un_informe_sin_filas_no_falla():
    hoja = load_workbook(construir_libro("Vacío", [col("Código")], [])).active
    assert hoja["A1"].value == "Almex Operaciones — Vacío"
    assert hoja.max_row == 4  # encabezado y títulos, sin datos ni totales


def test_los_valores_none_no_rompen_los_totales():
    columnas = [col("Nombre"), col("Monto", formato=CLP, total="suma")]
    hoja = load_workbook(construir_libro("Nulos", columnas, [["a", None], ["b", 500]])).active
    assert hoja.cell(row=hoja.max_row, column=2).value == 500


# --- Descargas desde las vistas ---


def test_descarga_de_stock_con_formato_correcto(client, empresa, usuario_admin, db):
    db.session.add(
        ItemConteoInventario(
            empresa_id=empresa.id, codigo="COD-001", nombre="Perno",
            cantidad_qms=10, cantidad_defontana=4,
        )
    )
    db.session.commit()
    login(client, "admin@test.cl")

    respuesta = client.get("/inventario/stock.xlsx")
    assert respuesta.status_code == 200
    assert "spreadsheetml" in respuesta.headers["Content-Type"]
    assert ".xlsx" in respuesta.headers["Content-Disposition"]

    hoja = _hoja(respuesta)
    assert hoja.cell(row=5, column=1).value == "COD-001"


def test_la_descarga_respeta_los_filtros_de_la_pantalla(client, empresa, usuario_admin, db):
    db.session.add_all([
        ItemConteoInventario(empresa_id=empresa.id, codigo="OK-1", nombre="Con stock", cantidad_qms=50, cantidad_defontana=50),
        ItemConteoInventario(empresa_id=empresa.id, codigo="DIF-1", nombre="Con diferencia", cantidad_qms=1, cantidad_defontana=99),
    ])
    db.session.commit()
    login(client, "admin@test.cl")

    hoja = _hoja(client.get("/inventario/stock.xlsx?filtro=diferencias"))
    codigos = [hoja.cell(row=fila, column=1).value for fila in range(5, hoja.max_row)]

    assert "DIF-1" in codigos
    assert "OK-1" not in codigos


def test_descarga_de_contratos(client, empresa, usuario_admin, db):
    cliente = Cliente(empresa_id=empresa.id, rut="11.111.111-1", razon_social="Cliente Uno")
    db.session.add(cliente)
    db.session.commit()
    db.session.add(
        ContratoCliente(
            empresa_id=empresa.id, cliente_id=cliente.id, numero_contrato="C-100",
            objeto="Mantención", fecha_inicio=date.today() - timedelta(days=5),
            fecha_termino=date.today() + timedelta(days=100), monto=2500000,
        )
    )
    db.session.commit()
    login(client, "admin@test.cl")

    hoja = _hoja(client.get("/contratos/lista.xlsx"))
    assert hoja.cell(row=5, column=1).value == "C-100"
    assert hoja.cell(row=5, column=2).value == "Cliente Uno"
    assert hoja.cell(row=5, column=9).value == 2500000


def test_descarga_de_activos_fijos(client, empresa, usuario_admin, db):
    db.session.add(
        ActivoFijo(
            empresa_id=empresa.id, codigo_activo="AF-1", nombre="Camioneta", categoria="Vehículos",
            fecha_compra=date.today() - timedelta(days=400), valor_compra=12000000, vida_util_meses=60,
        )
    )
    db.session.commit()
    login(client, "admin@test.cl")

    hoja = _hoja(client.get("/activos-fijos/lista.xlsx"))
    assert hoja.cell(row=5, column=1).value == "AF-1"
    assert hoja.cell(row=5, column=5).value == "En uso"
    assert hoja.cell(row=5, column=7).value == 12000000


def test_las_descargas_exigen_permiso(client, empresa, usuario_bodega):
    """Bodega ve inventario pero no contratos ni activos fijos."""
    login(client, "bodega@test.cl")

    assert client.get("/inventario/stock.xlsx").status_code == 200
    assert client.get("/contratos/lista.xlsx").status_code == 403
    assert client.get("/activos-fijos/lista.xlsx").status_code == 403


def test_las_descargas_exigen_sesion(client, empresa):
    respuesta = client.get("/inventario/stock.xlsx")
    assert respuesta.status_code == 302
    assert "/login" in respuesta.headers["Location"]
